"""Path C processor: hybrid splitter with enhanced region detection."""

from __future__ import annotations

import hashlib
import logging
import re
import time
import uuid
from pathlib import Path

import openpyxl
import pandas as pd

from ingestkit_excel.config import ExcelProcessorConfig
from ingestkit_excel.errors import ErrorCode, IngestError
from ingestkit_excel.models import (
    ChunkMetadata,
    ChunkPayload,
    ClassificationResult,
    ClassificationStageResult,
    EmbedStageResult,
    FileProfile,
    FileType,
    IngestionMethod,
    ParseStageResult,
    ProcessingResult,
    RegionType,
    SheetProfile,
    SheetRegion,
    WrittenArtifacts,
)
from ingestkit_excel.processors.structured_db import clean_name, deduplicate_names
from ingestkit_excel.protocols import (
    EmbeddingBackend,
    StructuredDBBackend,
    VectorStoreBackend,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_BLANK_ROW_THRESHOLD = 2
_BLANK_COL_THRESHOLD = 2
_FORMATTING_TRANSITION_WINDOW = 5
_NUMERIC_HEAVY_THRESHOLD = 0.6
_TEXT_HEAVY_THRESHOLD = 0.6
_HEADER_FOOTER_MAX_ROWS = 5
_MATRIX_MIN_HEADERS = 2


# ---------------------------------------------------------------------------
# Processor
# ---------------------------------------------------------------------------


class HybridSplitter:
    """Path C processor: detects regions within each sheet, classifies them
    independently, and processes Type A regions via DB ingestion and Type B
    regions via text serialization.

    The constructor accepts the structured processor and text serializer
    instances per the spec, but does NOT delegate to their ``process()``
    methods because sub-processors hardcode ``region_id=None``.  Instead,
    backends are extracted via private attributes and used directly.
    """

    def __init__(
        self,
        structured_processor: object,
        text_serializer: object,
        config: ExcelProcessorConfig,
    ) -> None:
        # Extract backends from the structured processor
        self._db: StructuredDBBackend = structured_processor._db  # type: ignore[attr-defined]
        self._vector_store: VectorStoreBackend = structured_processor._vector_store  # type: ignore[attr-defined]
        self._embedder: EmbeddingBackend = structured_processor._embedder  # type: ignore[attr-defined]
        self._config = config

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def process(
        self,
        file_path: str,
        profile: FileProfile,
        ingest_key: str,
        ingest_run_id: str,
        parse_result: ParseStageResult,
        classification_result: ClassificationStageResult,
        classification: ClassificationResult,
    ) -> ProcessingResult:
        """Process an Excel file via Path C (hybrid split).

        Iterates sheets, detects regions, classifies each region, processes
        Type A regions as tabular data and Type B regions as text, and merges
        all results into a single ProcessingResult.
        """
        start_time = time.monotonic()

        config = self._config
        collection = config.default_collection
        source_uri = f"file://{Path(file_path).resolve().as_posix()}"
        db_uri = self._db.get_connection_uri()

        errors: list[str] = []
        warnings: list[str] = []
        error_details: list[IngestError] = []

        written = WrittenArtifacts(vector_collection=collection)
        tables: list[str] = []
        total_chunks = 0
        total_texts_embedded = 0
        embed_duration = 0.0

        # Ensure vector collection exists
        vector_size = self._embedder.dimension()
        self._vector_store.ensure_collection(collection, vector_size)

        wb = openpyxl.load_workbook(file_path, data_only=True)
        chunk_index_counter = 0

        for sheet in profile.sheets:
            # --- Skip logic (identical to Path A/B) ---
            if sheet.is_hidden:
                warnings.append(ErrorCode.W_SHEET_SKIPPED_HIDDEN.value)
                logger.info("Skipping hidden sheet: %s", sheet.name)
                continue
            if sheet.row_count == 0 and sheet.col_count == 0:
                warnings.append(ErrorCode.W_SHEET_SKIPPED_CHART.value)
                logger.info("Skipping chart-only sheet: %s", sheet.name)
                continue
            if sheet.row_count > config.max_rows_in_memory:
                warnings.append(ErrorCode.W_ROWS_TRUNCATED.value)
                logger.warning(
                    "Sheet %s has %d rows, exceeds max_rows_in_memory (%d), skipping",
                    sheet.name,
                    sheet.row_count,
                    config.max_rows_in_memory,
                )
                continue

            try:
                ws = wb[sheet.name]

                # Compute all_rows ONCE per sheet (optimization note from PLAN-CHECK)
                all_rows: list[list] = [
                    [cell.value for cell in row] for row in ws.iter_rows()
                ]

                # Detect regions
                try:
                    regions = self._detect_regions(ws, sheet, all_rows)
                except Exception as exc:
                    error_code = ErrorCode.E_PROCESS_REGION_DETECT
                    errors.append(error_code.value)
                    error_details.append(
                        IngestError(
                            code=error_code,
                            message=str(exc),
                            sheet_name=sheet.name,
                            stage="process",
                            recoverable=False,
                        )
                    )
                    logger.exception(
                        "Region detection failed for sheet %s: %s",
                        sheet.name,
                        exc,
                    )
                    continue

                # Process each region
                for region in regions:
                    try:
                        if region.classified_as == FileType.TABULAR_DATA:
                            chunks, tbl_names, n_embedded, e_dur = (
                                self._process_type_a_region(
                                    file_path=file_path,
                                    region=region,
                                    all_rows=all_rows,
                                    sheet=sheet,
                                    source_uri=source_uri,
                                    db_uri=db_uri,
                                    collection=collection,
                                    ingest_key=ingest_key,
                                    ingest_run_id=ingest_run_id,
                                    chunk_index_counter=chunk_index_counter,
                                    tables=tables,
                                )
                            )
                            for c in chunks:
                                written.vector_point_ids.append(c.id)
                            tables.extend(tbl_names)
                            written.db_table_names.extend(tbl_names)
                            total_chunks += len(chunks)
                            total_texts_embedded += n_embedded
                            embed_duration += e_dur
                            chunk_index_counter += len(chunks)
                        else:
                            # Type B (FORMATTED_DOCUMENT or any non-tabular)
                            chunks, n_embedded, e_dur = (
                                self._process_type_b_region(
                                    region=region,
                                    all_rows=all_rows,
                                    sheet=sheet,
                                    source_uri=source_uri,
                                    collection=collection,
                                    ingest_key=ingest_key,
                                    ingest_run_id=ingest_run_id,
                                    chunk_index_counter=chunk_index_counter,
                                )
                            )
                            for c in chunks:
                                written.vector_point_ids.append(c.id)
                            total_chunks += len(chunks)
                            total_texts_embedded += n_embedded
                            embed_duration += e_dur
                            chunk_index_counter += len(chunks)

                    except Exception as exc:
                        error_code = self._classify_backend_error(exc)
                        errors.append(error_code.value)
                        error_details.append(
                            IngestError(
                                code=error_code,
                                message=str(exc),
                                sheet_name=sheet.name,
                                stage="process",
                                recoverable=False,
                            )
                        )
                        logger.exception(
                            "Error processing region %s in sheet %s: %s",
                            region.region_id,
                            sheet.name,
                            exc,
                        )
                        continue

            except Exception as exc:
                error_code = self._classify_backend_error(exc)
                errors.append(error_code.value)
                error_details.append(
                    IngestError(
                        code=error_code,
                        message=str(exc),
                        sheet_name=sheet.name,
                        stage="process",
                        recoverable=False,
                    )
                )
                logger.exception(
                    "Error processing sheet %s: %s", sheet.name, exc
                )
                continue

        wb.close()

        # --- Assemble EmbedStageResult ---
        embed_result = None
        if total_texts_embedded > 0:
            embed_result = EmbedStageResult(
                texts_embedded=total_texts_embedded,
                embedding_dimension=self._embedder.dimension(),
                embed_duration_seconds=embed_duration,
            )

        elapsed = time.monotonic() - start_time

        return ProcessingResult(
            file_path=file_path,
            ingest_key=ingest_key,
            ingest_run_id=ingest_run_id,
            tenant_id=config.tenant_id,
            parse_result=parse_result,
            classification_result=classification_result,
            embed_result=embed_result,
            classification=classification,
            ingestion_method=IngestionMethod.HYBRID_SPLIT,
            chunks_created=total_chunks,
            tables_created=len(tables),
            tables=tables,
            written=written,
            errors=errors,
            warnings=warnings,
            error_details=error_details,
            processing_time_seconds=elapsed,
        )

    # ------------------------------------------------------------------
    # Region detection
    # ------------------------------------------------------------------

    def _detect_regions(
        self,
        ws,
        sheet: SheetProfile,
        all_rows: list[list],
    ) -> list[SheetRegion]:
        """Orchestrate the 5 heuristics to detect regions in a sheet.

        Returns a list of SheetRegion objects with classified_as set.
        """
        if not all_rows:
            return []

        total_rows = len(all_rows)
        col_count = max((len(r) for r in all_rows), default=0)

        # 1. Detect header/footer
        header_region, footer_region = self._detect_header_footer(
            ws, all_rows, total_rows
        )

        # 2. Find blank row boundaries
        blank_row_boundaries = self._detect_blank_row_boundaries(all_rows)

        # 3. Find blank col boundaries
        blank_col_boundaries = self._detect_blank_col_boundaries(
            all_rows, col_count
        )

        # 4. Detect merged blocks
        merged_regions = self._detect_merged_blocks(ws, sheet.name)

        # 5. Detect formatting transitions
        formatting_boundaries = self._detect_formatting_transitions(all_rows)

        # Combine all row-based boundaries
        all_row_boundaries = sorted(
            set(blank_row_boundaries) | set(formatting_boundaries)
        )

        # Determine content area (excluding header/footer)
        content_start = 0
        content_end = total_rows
        if header_region is not None:
            content_start = header_region.end_row
        if footer_region is not None:
            content_end = footer_region.start_row

        # Build regions from boundaries within the content area
        regions: list[SheetRegion] = []

        # Add header region if found
        if header_region is not None:
            header_region.classified_as = FileType.FORMATTED_DOCUMENT
            regions.append(header_region)

        # Add merged block regions
        for mr in merged_regions:
            mr.classified_as = FileType.FORMATTED_DOCUMENT
            regions.append(mr)

        # Split content area by row boundaries
        effective_boundaries = [
            b for b in all_row_boundaries
            if content_start < b < content_end
        ]

        # Build sub-regions from boundary splits
        split_points = [content_start] + effective_boundaries + [content_end]
        region_idx = len(regions)

        for i in range(len(split_points) - 1):
            start = split_points[i]
            end = split_points[i + 1]

            if end <= start:
                continue

            # Check if region is all blank
            region_rows = all_rows[start:end]
            if all(
                all(v is None or str(v).strip() == "" for v in row)
                for row in region_rows
            ):
                continue

            # Determine column bounds, respecting column boundaries
            start_col = 0
            end_col = col_count

            # Check for matrix regions
            is_matrix = self._detect_matrix_regions(
                all_rows, start, end, start_col, end_col
            )

            region_type = RegionType.MATRIX_BLOCK if is_matrix else RegionType.DATA_TABLE
            region_id = f"{sheet.name}_r{region_idx}"

            sr = SheetRegion(
                sheet_name=sheet.name,
                region_id=region_id,
                start_row=start,
                end_row=end,
                start_col=start_col,
                end_col=end_col,
                region_type=region_type,
                detection_confidence=0.8,
            )
            sr.classified_as = self._classify_region(all_rows, sr)
            regions.append(sr)
            region_idx += 1

        # Add footer region if found
        if footer_region is not None:
            footer_region.classified_as = FileType.FORMATTED_DOCUMENT
            regions.append(footer_region)

        # If no regions detected from boundaries, create one region for
        # the entire content area
        if not regions and total_rows > 0:
            sr = SheetRegion(
                sheet_name=sheet.name,
                region_id=f"{sheet.name}_r0",
                start_row=0,
                end_row=total_rows,
                start_col=0,
                end_col=col_count,
                region_type=RegionType.DATA_TABLE,
                detection_confidence=0.7,
            )
            sr.classified_as = self._classify_region(all_rows, sr)
            regions.append(sr)

        return regions

    @staticmethod
    def _detect_blank_row_boundaries(all_rows: list[list]) -> list[int]:
        """Find positions where >= _BLANK_ROW_THRESHOLD consecutive blank rows occur.

        Returns the row index AFTER the blank gap (i.e. where the next region starts).
        """
        boundaries: list[int] = []
        consecutive_blank = 0
        gap_start = -1

        for idx, row in enumerate(all_rows):
            is_blank = all(v is None or str(v).strip() == "" for v in row)
            if is_blank:
                if consecutive_blank == 0:
                    gap_start = idx
                consecutive_blank += 1
            else:
                if consecutive_blank >= _BLANK_ROW_THRESHOLD:
                    boundaries.append(gap_start)
                consecutive_blank = 0

        # Handle trailing blank rows
        if consecutive_blank >= _BLANK_ROW_THRESHOLD:
            boundaries.append(gap_start)

        return boundaries

    @staticmethod
    def _detect_blank_col_boundaries(
        all_rows: list[list], col_count: int
    ) -> list[int]:
        """Find columns where >= _BLANK_COL_THRESHOLD consecutive columns are all blank.

        Returns column indices where blank gaps start.
        """
        if col_count == 0 or not all_rows:
            return []

        boundaries: list[int] = []
        consecutive_blank = 0
        gap_start = -1

        for col_idx in range(col_count):
            is_blank = all(
                (col_idx >= len(row) or row[col_idx] is None or str(row[col_idx]).strip() == "")
                for row in all_rows
            )
            if is_blank:
                if consecutive_blank == 0:
                    gap_start = col_idx
                consecutive_blank += 1
            else:
                if consecutive_blank >= _BLANK_COL_THRESHOLD:
                    boundaries.append(gap_start)
                consecutive_blank = 0

        if consecutive_blank >= _BLANK_COL_THRESHOLD:
            boundaries.append(gap_start)

        return boundaries

    @staticmethod
    def _detect_merged_blocks(ws, sheet_name: str) -> list[SheetRegion]:
        """Detect large merged cell regions and return them as HEADER_BLOCK regions."""
        regions: list[SheetRegion] = []
        for idx, mr in enumerate(ws.merged_cells.ranges):
            span_cols = mr.max_col - mr.min_col + 1
            if span_cols >= 2:
                regions.append(
                    SheetRegion(
                        sheet_name=sheet_name,
                        region_id=f"{sheet_name}_merged_{idx}",
                        start_row=mr.min_row - 1,  # Convert to 0-based
                        end_row=mr.max_row,  # 0-based exclusive
                        start_col=mr.min_col - 1,  # 0-based
                        end_col=mr.max_col,  # 0-based exclusive
                        region_type=RegionType.HEADER_BLOCK,
                        detection_confidence=0.9,
                    )
                )
        return regions

    @staticmethod
    def _detect_formatting_transitions(all_rows: list[list]) -> list[int]:
        """Detect boundaries where the data type ratio shifts significantly.

        Uses a sliding window to detect numeric-heavy vs text-heavy transitions.
        """
        if len(all_rows) < _FORMATTING_TRANSITION_WINDOW * 2:
            return []

        boundaries: list[int] = []

        def _compute_numeric_ratio(rows: list[list]) -> float:
            total = 0
            numeric = 0
            for row in rows:
                for val in row:
                    if val is not None and str(val).strip():
                        total += 1
                        try:
                            float(str(val))
                            numeric += 1
                        except (ValueError, TypeError):
                            pass
            return numeric / total if total > 0 else 0.0

        window = _FORMATTING_TRANSITION_WINDOW
        for i in range(window, len(all_rows) - window + 1):
            before = _compute_numeric_ratio(all_rows[i - window : i])
            after = _compute_numeric_ratio(all_rows[i : i + window])

            # Detect significant transition
            before_numeric = before >= _NUMERIC_HEAVY_THRESHOLD
            after_numeric = after >= _NUMERIC_HEAVY_THRESHOLD
            before_text = before <= (1 - _TEXT_HEAVY_THRESHOLD)
            after_text = after <= (1 - _TEXT_HEAVY_THRESHOLD)

            if (before_numeric and after_text) or (before_text and after_numeric):
                boundaries.append(i)

        return boundaries

    @staticmethod
    def _detect_header_footer(
        ws,
        all_rows: list[list],
        total_rows: int,
    ) -> tuple[SheetRegion | None, SheetRegion | None]:
        """Check first/last N rows for full-width merges indicating header/footer."""
        header_region: SheetRegion | None = None
        footer_region: SheetRegion | None = None

        if total_rows == 0:
            return None, None

        # Check for header in first N rows
        check_rows = min(_HEADER_FOOTER_MAX_ROWS, total_rows)
        for mr in ws.merged_cells.ranges:
            span_cols = mr.max_col - mr.min_col + 1
            if span_cols >= 3 and mr.min_row <= check_rows:
                val = ws.cell(mr.min_row, mr.min_col).value
                if val is not None and str(val).strip():
                    header_region = SheetRegion(
                        sheet_name=ws.title if hasattr(ws, "title") else "Sheet",
                        region_id="header",
                        start_row=0,
                        end_row=mr.max_row,
                        start_col=0,
                        end_col=mr.max_col,
                        region_type=RegionType.HEADER_BLOCK,
                        detection_confidence=0.9,
                    )
                    break

        # Check for footer in last N rows
        footer_start = max(0, total_rows - check_rows)
        for mr in ws.merged_cells.ranges:
            span_cols = mr.max_col - mr.min_col + 1
            if span_cols >= 3 and mr.min_row > footer_start:
                val = ws.cell(mr.min_row, mr.min_col).value
                if val is not None and str(val).strip():
                    footer_region = SheetRegion(
                        sheet_name=ws.title if hasattr(ws, "title") else "Sheet",
                        region_id="footer",
                        start_row=mr.min_row - 1,
                        end_row=total_rows,
                        start_col=0,
                        end_col=mr.max_col,
                        region_type=RegionType.FOOTER_BLOCK,
                        detection_confidence=0.85,
                    )
                    break

        return header_region, footer_region

    @staticmethod
    def _detect_matrix_regions(
        all_rows: list[list],
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
    ) -> bool:
        """Check if a region is a matrix (row headers in col 0, col headers in row 0,
        corner cell empty).
        """
        region_rows = all_rows[start_row:end_row]
        if len(region_rows) < 2:
            return False

        first_row = region_rows[0]
        if len(first_row) < 2:
            return False

        # Corner should be empty
        corner = first_row[start_col] if start_col < len(first_row) else None
        if corner is not None and str(corner).strip():
            return False

        # Check column headers (row 0, cols 1+)
        col_headers = 0
        for c in range(start_col + 1, min(end_col, len(first_row))):
            if first_row[c] is not None and str(first_row[c]).strip():
                col_headers += 1

        if col_headers < _MATRIX_MIN_HEADERS:
            return False

        # Check row headers (col 0, rows 1+)
        row_headers = 0
        for row in region_rows[1:]:
            if start_col < len(row) and row[start_col] is not None and str(row[start_col]).strip():
                row_headers += 1

        return row_headers > 0

    # ------------------------------------------------------------------
    # Region classification
    # ------------------------------------------------------------------

    def _classify_region(
        self, all_rows: list[list], region: SheetRegion
    ) -> FileType:
        """Classify a region using Tier 1 signals on region data.

        >= 4 Type A signals -> TABULAR_DATA, else FORMATTED_DOCUMENT.
        """
        signals = self._compute_region_signals(all_rows, region)
        type_a_count = sum(1 for v in signals.values() if v)

        if type_a_count >= self._config.tier1_high_confidence_signals:
            return FileType.TABULAR_DATA
        return FileType.FORMATTED_DOCUMENT

    def _compute_region_signals(
        self, all_rows: list[list], region: SheetRegion
    ) -> dict[str, bool]:
        """Compute the 5 Tier 1 binary signals for a region."""
        cfg = self._config
        region_rows = all_rows[region.start_row : region.end_row]
        row_count = len(region_rows)

        # 1. row_count
        sig_row_count = row_count >= cfg.min_row_count_for_tabular

        # 2. merged_cell_ratio (region has no merged tracking, approximate as low)
        # For region-level, we default to True (low merged ratio) since we
        # already extracted merged blocks separately
        sig_merged_cell_ratio = region.region_type != RegionType.HEADER_BLOCK

        # 3. column_type_consistency
        if row_count > 1:
            col_count = max((len(r) for r in region_rows), default=0)
            consistent_cols = 0
            for c in range(col_count):
                types_seen: set[str] = set()
                for row in region_rows:
                    if c < len(row) and row[c] is not None:
                        val = row[c]
                        try:
                            float(str(val))
                            types_seen.add("numeric")
                        except (ValueError, TypeError):
                            types_seen.add("text")
                if len(types_seen) <= 1:
                    consistent_cols += 1
            consistency = consistent_cols / col_count if col_count > 0 else 0.0
        else:
            consistency = 0.0
        sig_column_consistency = consistency >= cfg.column_consistency_threshold

        # 4. header_detected
        sig_header = False
        if region_rows:
            first_row = region_rows[0]
            non_empty = [v for v in first_row if v is not None and str(v).strip()]
            if non_empty:
                all_numeric = all(
                    str(v).replace(".", "").replace("-", "").isdigit()
                    for v in non_empty
                )
                distinct = len(set(str(v).strip().lower() for v in non_empty)) == len(non_empty)
                sig_header = not all_numeric and distinct

        # 5. numeric_ratio
        total_cells = 0
        numeric_cells = 0
        for row in region_rows:
            for val in row:
                if val is not None and str(val).strip():
                    total_cells += 1
                    try:
                        float(str(val))
                        numeric_cells += 1
                    except (ValueError, TypeError):
                        pass
        numeric_ratio = numeric_cells / total_cells if total_cells > 0 else 0.0
        sig_numeric = numeric_ratio >= cfg.numeric_ratio_threshold

        return {
            "row_count": sig_row_count,
            "merged_cell_ratio": sig_merged_cell_ratio,
            "column_type_consistency": sig_column_consistency,
            "header_detected": sig_header,
            "numeric_ratio": sig_numeric,
        }

    # ------------------------------------------------------------------
    # Type A region processing
    # ------------------------------------------------------------------

    def _process_type_a_region(
        self,
        *,
        file_path: str,
        region: SheetRegion,
        all_rows: list[list],
        sheet: SheetProfile,
        source_uri: str,
        db_uri: str,
        collection: str,
        ingest_key: str,
        ingest_run_id: str,
        chunk_index_counter: int,
        tables: list[str],
    ) -> tuple[list[ChunkPayload], list[str], int, float]:
        """Process a Type A (tabular) region.

        Returns (chunks, table_names, texts_embedded, embed_duration).
        """
        config = self._config
        embed_duration = 0.0
        texts_embedded = 0

        # Extract DataFrame from region
        region_rows = all_rows[region.start_row : region.end_row]
        if not region_rows:
            return [], [], 0, 0.0

        # Build DataFrame from region data
        col_count = max((len(r) for r in region_rows), default=0)
        # Pad rows to uniform length
        padded = [
            list(r) + [None] * (col_count - len(r)) for r in region_rows
        ]

        # Use first row as header if it looks like one
        first_row = padded[0] if padded else []
        non_empty_headers = [v for v in first_row if v is not None and str(v).strip()]
        use_header = bool(non_empty_headers) and not all(
            str(v).replace(".", "").replace("-", "").isdigit()
            for v in non_empty_headers
        )

        if use_header and len(padded) > 1:
            columns = [str(v) if v is not None else f"column_{i}" for i, v in enumerate(first_row)]
            data_rows = padded[1:]
        else:
            columns = [f"column_{i}" for i in range(col_count)]
            data_rows = padded

        df = pd.DataFrame(data_rows, columns=columns)

        # Clean column names
        if config.clean_column_names:
            cleaned = [clean_name(str(c)) for c in df.columns]
            cleaned = deduplicate_names(cleaned)
            df.columns = cleaned

        # Create table
        table_name = clean_name(f"{sheet.name}_{region.region_id}")
        if not table_name:
            table_name = f"region_{region.region_id}"
        # Deduplicate table names
        if table_name in tables:
            suffix = 1
            while f"{table_name}_{suffix}" in tables:
                suffix += 1
            table_name = f"{table_name}_{suffix}"

        self._db.create_table_from_dataframe(table_name, df)
        new_tables = [table_name]

        # Generate schema description
        schema_text = self._generate_schema_description(table_name, df)

        # Embed schema
        embed_start = time.monotonic()
        vectors = self._embedder.embed(
            [schema_text], timeout=config.backend_timeout_seconds
        )
        embed_duration += time.monotonic() - embed_start
        texts_embedded += 1

        chunk_hash = hashlib.sha256(schema_text.encode()).hexdigest()
        chunk_id = str(
            uuid.uuid5(uuid.NAMESPACE_URL, f"{ingest_key}:{chunk_hash}")
        )
        columns_list = list(df.columns)

        metadata = ChunkMetadata(
            source_uri=source_uri,
            source_format="xlsx",
            sheet_name=sheet.name,
            region_id=region.region_id,
            ingestion_method=IngestionMethod.HYBRID_SPLIT.value,
            parser_used=sheet.parser_used.value,
            parser_version=config.parser_version,
            chunk_index=chunk_index_counter,
            chunk_hash=chunk_hash,
            ingest_key=ingest_key,
            ingest_run_id=ingest_run_id,
            tenant_id=config.tenant_id,
            table_name=table_name,
            db_uri=db_uri,
            row_count=len(df),
            columns=columns_list,
        )
        chunk = ChunkPayload(
            id=chunk_id,
            text=schema_text,
            vector=vectors[0],
            metadata=metadata,
        )

        self._vector_store.upsert_chunks(collection, [chunk])
        return [chunk], new_tables, texts_embedded, embed_duration

    # ------------------------------------------------------------------
    # Type B region processing
    # ------------------------------------------------------------------

    def _process_type_b_region(
        self,
        *,
        region: SheetRegion,
        all_rows: list[list],
        sheet: SheetProfile,
        source_uri: str,
        collection: str,
        ingest_key: str,
        ingest_run_id: str,
        chunk_index_counter: int,
    ) -> tuple[list[ChunkPayload], int, float]:
        """Process a Type B (text) region.

        Returns (chunks, texts_embedded, embed_duration).
        """
        config = self._config
        embed_duration = 0.0
        texts_embedded = 0

        region_rows = all_rows[region.start_row : region.end_row]
        if not region_rows:
            return [], 0, 0.0

        # Serialize content into natural language text
        text = self._serialize_region_text(region_rows, region)
        if not text.strip():
            return [], 0, 0.0

        chunk_hash = hashlib.sha256(text.encode()).hexdigest()
        chunk_id = str(
            uuid.uuid5(uuid.NAMESPACE_URL, f"{ingest_key}:{chunk_hash}")
        )

        # Embed
        embed_start = time.monotonic()
        vectors = self._embedder.embed(
            [text], timeout=config.backend_timeout_seconds
        )
        embed_duration += time.monotonic() - embed_start
        texts_embedded += 1

        metadata = ChunkMetadata(
            source_uri=source_uri,
            source_format="xlsx",
            sheet_name=sheet.name,
            region_id=region.region_id,
            ingestion_method=IngestionMethod.HYBRID_SPLIT.value,
            parser_used=sheet.parser_used.value,
            parser_version=config.parser_version,
            chunk_index=chunk_index_counter,
            chunk_hash=chunk_hash,
            ingest_key=ingest_key,
            ingest_run_id=ingest_run_id,
            tenant_id=config.tenant_id,
            table_name=None,
            db_uri=None,
            row_count=None,
            columns=None,
            original_structure=region.region_type.value,
        )
        chunk = ChunkPayload(
            id=chunk_id,
            text=text,
            vector=vectors[0],
            metadata=metadata,
        )

        self._vector_store.upsert_chunks(collection, [chunk])
        return [chunk], texts_embedded, embed_duration

    # ------------------------------------------------------------------
    # Helper methods
    # ------------------------------------------------------------------

    @staticmethod
    def _generate_schema_description(
        table_name: str, df: pd.DataFrame
    ) -> str:
        """Generate a natural language schema description for embedding."""
        lines = [f'Table "{table_name}" contains {len(df)} rows with columns:']
        for col in df.columns:
            series = df[col].dropna()
            if len(series) == 0:
                lines.append(f"- {col}: no data")
                continue
            # Infer type
            if pd.api.types.is_integer_dtype(series):
                lines.append(f"- {col} (integer): range {series.min()} to {series.max()}")
            elif pd.api.types.is_float_dtype(series):
                lines.append(f"- {col} (float): range {series.min()} to {series.max()}")
            elif pd.api.types.is_bool_dtype(series):
                lines.append(f"- {col} (boolean): true/false")
            else:
                n_unique = series.nunique()
                if n_unique < 20:
                    vals = ", ".join(str(v) for v in series.unique()[:20])
                    lines.append(f"- {col} (text): one of {vals}")
                else:
                    lines.append(f"- {col} (text): {n_unique} unique values")
        return "\n".join(lines)

    @staticmethod
    def _serialize_region_text(
        region_rows: list[list], region: SheetRegion
    ) -> str:
        """Serialize a text region's content into natural language."""
        paragraphs: list[str] = []
        for row in region_rows:
            cells = [str(v) for v in row if v is not None and str(v).strip()]
            if cells:
                paragraphs.append(" ".join(cells))
        return "\n\n".join(paragraphs)

    @staticmethod
    def _classify_backend_error(exc: Exception) -> ErrorCode:
        """Map an exception to the most appropriate ErrorCode."""
        msg = str(exc).lower()
        if "timeout" in msg or "timed out" in msg:
            if "embed" in msg:
                return ErrorCode.E_BACKEND_EMBED_TIMEOUT
            if "vector" in msg or "qdrant" in msg or "collection" in msg:
                return ErrorCode.E_BACKEND_VECTOR_TIMEOUT
            if "db" in msg or "database" in msg or "sql" in msg:
                return ErrorCode.E_BACKEND_DB_TIMEOUT
            return ErrorCode.E_BACKEND_DB_TIMEOUT
        if "connect" in msg or "connection" in msg:
            if "embed" in msg:
                return ErrorCode.E_BACKEND_EMBED_CONNECT
            if "vector" in msg or "qdrant" in msg or "collection" in msg:
                return ErrorCode.E_BACKEND_VECTOR_CONNECT
            if "db" in msg or "database" in msg or "sql" in msg:
                return ErrorCode.E_BACKEND_DB_CONNECT
            return ErrorCode.E_BACKEND_DB_CONNECT
        # Default to region detect error for unknown exceptions
        return ErrorCode.E_PROCESS_REGION_DETECT

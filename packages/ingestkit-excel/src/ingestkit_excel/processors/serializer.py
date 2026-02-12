"""Path B processor: text serialization for document-formatted Excel files."""

from __future__ import annotations

import hashlib
import logging
import time
import uuid
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from ingestkit_excel.config import ExcelProcessorConfig
from ingestkit_excel.errors import ErrorCode, IngestError
from ingestkit_excel.models import (
    ChunkMetadata,
    ChunkPayload,
    ClassificationResult,
    ClassificationStageResult,
    EmbedStageResult,
    FileProfile,
    IngestionMethod,
    ParseStageResult,
    ProcessingResult,
    SheetProfile,
    WrittenArtifacts,
)
from ingestkit_excel.protocols import (
    EmbeddingBackend,
    VectorStoreBackend,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Internal data structures
# ---------------------------------------------------------------------------


@dataclass
class Section:
    """A detected logical section within a worksheet."""

    title: str
    sub_structure: str  # "table", "checklist", "matrix", "free_text"
    rows: list[list]  # raw cell values, one list per row
    start_row: int  # 1-based
    end_row: int  # 1-based
    col_count: int
    header_row: list[str] | None = None


# Status keywords used for checklist detection
_STATUS_KEYWORDS = frozenset({
    "status", "done", "complete", "pending", "checked", "yes", "no",
})


# ---------------------------------------------------------------------------
# Processor
# ---------------------------------------------------------------------------


class TextSerializer:
    """Path B processor: detects logical sections in document-formatted Excel
    files, serializes each section into natural language chunks, embeds them,
    and upserts to a vector store.
    """

    def __init__(
        self,
        vector_store: VectorStoreBackend,
        embedder: EmbeddingBackend,
        config: ExcelProcessorConfig,
    ) -> None:
        self._vector_store = vector_store
        self._embedder = embedder
        self._config = config

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def process(
        self,
        file_path: str,
        profile: FileProfile,
        ingest_key: str,
        ingest_run_id: str,
        parse_result: ParseStageResult,
        classification_result: ClassificationStageResult,
        classification: ClassificationResult,
    ) -> ProcessingResult:
        """Process an Excel file via Path B (text serialization).

        Args:
            file_path: Absolute path to the Excel file on disk.
            profile: Pre-computed structural profile of the file.
            ingest_key: Deterministic SHA-256 hex string.
            ingest_run_id: UUID4 string unique to this processing run.
            parse_result: Typed output of the parsing stage.
            classification_result: Typed output of the classification stage.
            classification: Simplified classification result.

        Returns:
            A fully-assembled ``ProcessingResult``.
        """
        start_time = time.monotonic()

        config = self._config
        collection = config.default_collection
        source_uri = f"file://{Path(file_path).resolve().as_posix()}"

        errors: list[str] = []
        warnings: list[str] = []
        error_details: list[IngestError] = []

        written = WrittenArtifacts(vector_collection=collection)
        total_chunks = 0
        total_texts_embedded = 0
        embed_duration = 0.0

        # Ensure vector collection exists
        vector_size = self._embedder.dimension()
        self._vector_store.ensure_collection(collection, vector_size)

        wb = openpyxl.load_workbook(file_path, data_only=True)
        chunk_index_counter = 0  # global chunk index across all sheets

        for sheet in profile.sheets:
            # --- Skip logic (identical to Path A) ---
            if sheet.is_hidden:
                warnings.append(ErrorCode.W_SHEET_SKIPPED_HIDDEN.value)
                logger.info("Skipping hidden sheet: %s", sheet.name)
                continue
            if sheet.row_count == 0 and sheet.col_count == 0:
                warnings.append(ErrorCode.W_SHEET_SKIPPED_CHART.value)
                logger.info("Skipping chart-only sheet: %s", sheet.name)
                continue
            if sheet.row_count > config.max_rows_in_memory:
                warnings.append(ErrorCode.W_ROWS_TRUNCATED.value)
                logger.warning(
                    "Sheet %s has %d rows, exceeds max_rows_in_memory (%d), skipping",
                    sheet.name,
                    sheet.row_count,
                    config.max_rows_in_memory,
                )
                continue

            try:
                ws = wb[sheet.name]
                sections = self._detect_sections(ws, sheet)

                # Build chunks for this sheet
                sheet_chunks: list[ChunkPayload] = []
                for section in sections:
                    text = self._serialize_section(section)
                    if not text.strip():
                        continue

                    chunk_hash = hashlib.sha256(text.encode()).hexdigest()
                    chunk_id = str(
                        uuid.uuid5(
                            uuid.NAMESPACE_URL,
                            f"{ingest_key}:{chunk_hash}",
                        )
                    )

                    metadata = ChunkMetadata(
                        source_uri=source_uri,
                        source_format="xlsx",
                        sheet_name=sheet.name,
                        region_id=None,
                        ingestion_method=IngestionMethod.TEXT_SERIALIZATION.value,
                        parser_used=sheet.parser_used.value,
                        parser_version=config.parser_version,
                        chunk_index=chunk_index_counter,
                        chunk_hash=chunk_hash,
                        ingest_key=ingest_key,
                        ingest_run_id=ingest_run_id,
                        tenant_id=config.tenant_id,
                        table_name=None,
                        db_uri=None,
                        row_count=None,
                        columns=None,
                        section_title=section.title,
                        original_structure=section.sub_structure,
                    )
                    sheet_chunks.append(
                        ChunkPayload(
                            id=chunk_id,
                            text=text,
                            vector=[],
                            metadata=metadata,
                        )
                    )
                    chunk_index_counter += 1

                # Embed in batches
                for batch_start in range(
                    0, len(sheet_chunks), config.embedding_batch_size
                ):
                    batch = sheet_chunks[
                        batch_start : batch_start + config.embedding_batch_size
                    ]
                    texts = [c.text for c in batch]
                    embed_start = time.monotonic()
                    vectors = self._embedder.embed(
                        texts,
                        timeout=config.backend_timeout_seconds,
                    )
                    embed_duration += time.monotonic() - embed_start
                    total_texts_embedded += len(texts)
                    for c, vec in zip(batch, vectors):
                        c.vector = vec

                    self._vector_store.upsert_chunks(collection, list(batch))
                    for c in batch:
                        written.vector_point_ids.append(c.id)

                total_chunks += len(sheet_chunks)

            except Exception as exc:
                error_code = self._classify_backend_error(exc)
                errors.append(error_code.value)
                error_details.append(
                    IngestError(
                        code=error_code,
                        message=str(exc),
                        sheet_name=sheet.name,
                        stage="process",
                        recoverable=False,
                    )
                )
                logger.exception(
                    "Error processing sheet %s: %s", sheet.name, exc
                )
                continue

        wb.close()

        # --- Assemble EmbedStageResult ---
        embed_result = None
        if total_texts_embedded > 0:
            embed_result = EmbedStageResult(
                texts_embedded=total_texts_embedded,
                embedding_dimension=self._embedder.dimension(),
                embed_duration_seconds=embed_duration,
            )

        elapsed = time.monotonic() - start_time

        return ProcessingResult(
            file_path=file_path,
            ingest_key=ingest_key,
            ingest_run_id=ingest_run_id,
            tenant_id=config.tenant_id,
            parse_result=parse_result,
            classification_result=classification_result,
            embed_result=embed_result,
            classification=classification,
            ingestion_method=IngestionMethod.TEXT_SERIALIZATION,
            chunks_created=total_chunks,
            tables_created=0,
            tables=[],
            written=written,
            errors=errors,
            warnings=warnings,
            error_details=error_details,
            processing_time_seconds=elapsed,
        )

    # ------------------------------------------------------------------
    # Section detection
    # ------------------------------------------------------------------

    @staticmethod
    def _detect_sections(ws, sheet: SheetProfile) -> list[Section]:
        """Detect logical sections in a worksheet.

        Uses merged header rows and blank row separators to split the
        worksheet into coherent sections, then classifies each section's
        sub-structure.
        """
        all_rows: list[list] = [
            [cell.value for cell in row] for row in ws.iter_rows()
        ]
        if not all_rows:
            return []

        # Build merged header map: row_index -> (value, span_width)
        merged_headers: dict[int, tuple[str, int]] = {}
        for mr in ws.merged_cells.ranges:
            span_width = mr.max_col - mr.min_col + 1
            if span_width >= 2:
                val = ws.cell(mr.min_row, mr.min_col).value
                if val is not None and str(val).strip():
                    # Convert to 0-based index
                    merged_headers[mr.min_row - 1] = (str(val).strip(), span_width)

        # Scan rows to find boundaries
        def _is_blank_row(row: list) -> bool:
            return all(v is None or str(v).strip() == "" for v in row)

        sections: list[Section] = []
        current_rows: list[list] = []
        current_title: str | None = None
        current_start: int = 1  # 1-based
        section_counter = 0

        for idx, row in enumerate(all_rows):
            # Check if this is a merged header row
            if idx in merged_headers:
                # Flush current section if it has data
                if current_rows:
                    section_counter += 1
                    title = current_title or f"Section {section_counter}"
                    col_count = max((len(r) for r in current_rows), default=0)
                    sec = Section(
                        title=title,
                        sub_structure="free_text",
                        rows=current_rows,
                        start_row=current_start,
                        end_row=current_start + len(current_rows) - 1,
                        col_count=col_count,
                    )
                    TextSerializer._classify_sub_structure(sec)
                    sections.append(sec)
                    current_rows = []

                current_title = merged_headers[idx][0]
                current_start = idx + 1  # 1-based
                continue

            if _is_blank_row(row):
                # Blank row: flush if we have accumulated data
                if current_rows:
                    section_counter += 1
                    title = current_title or f"Section {section_counter}"
                    col_count = max((len(r) for r in current_rows), default=0)
                    sec = Section(
                        title=title,
                        sub_structure="free_text",
                        rows=current_rows,
                        start_row=current_start,
                        end_row=current_start + len(current_rows) - 1,
                        col_count=col_count,
                    )
                    TextSerializer._classify_sub_structure(sec)
                    sections.append(sec)
                    current_rows = []
                    current_title = None
                continue

            # Data row
            if not current_rows and current_title is None:
                current_start = idx + 1  # 1-based
            current_rows.append(row)

        # Flush remaining rows
        if current_rows:
            section_counter += 1
            title = current_title or f"Section {section_counter}"
            col_count = max((len(r) for r in current_rows), default=0)
            sec = Section(
                title=title,
                sub_structure="free_text",
                rows=current_rows,
                start_row=current_start,
                end_row=current_start + len(current_rows) - 1,
                col_count=col_count,
            )
            TextSerializer._classify_sub_structure(sec)
            sections.append(sec)

        # If no boundaries found and the section has only a fallback title, use sheet name.
        # Check for exact "Section N" pattern (auto-generated) vs user content.
        if (
            len(sections) == 1
            and sections[0].title.startswith("Section ")
            and sections[0].title[len("Section "):].isdigit()
        ):
            sections[0].title = sheet.name

        return sections

    # ------------------------------------------------------------------
    # Sub-structure classification
    # ------------------------------------------------------------------

    @staticmethod
    def _classify_sub_structure(section: Section) -> None:
        """Classify a section's sub-structure and set header_row if applicable.

        Mutates ``section.sub_structure`` and optionally ``section.header_row``.
        """
        rows = section.rows
        if len(rows) < 2:
            section.sub_structure = "free_text"
            return

        first_row = rows[0]
        first_row_strs = [str(v).strip().lower() if v is not None else "" for v in first_row]

        # --- Checklist detection (highest priority) ---
        has_status_col = any(val in _STATUS_KEYWORDS for val in first_row_strs)
        if has_status_col and len(rows) >= 2:
            section.sub_structure = "checklist"
            section.header_row = [str(v) if v is not None else "" for v in first_row]
            return

        # --- Matrix detection ---
        # Col 0 has non-empty values in data rows (row headers),
        # row 0 has non-empty values in cols 1+ (column headers),
        # and row0[0] is empty or a generic label (distinguishes from table).
        col0_populated = sum(
            1 for r in rows[1:] if r[0] is not None and str(r[0]).strip()
        )
        row0_cols_populated = sum(
            1 for v in first_row[1:] if v is not None and str(v).strip()
        )
        corner_empty = first_row[0] is None or str(first_row[0]).strip() == ""
        if (
            len(first_row) >= 2
            and corner_empty
            and col0_populated > 0
            and row0_cols_populated >= 2
            and col0_populated >= len(rows[1:]) * 0.5
        ):
            # Check intersection cells are populated
            intersect_populated = 0
            intersect_total = 0
            for r in rows[1:]:
                for v in r[1:]:
                    intersect_total += 1
                    if v is not None and str(v).strip():
                        intersect_populated += 1
            if intersect_total > 0 and intersect_populated / intersect_total > 0.3:
                section.sub_structure = "matrix"
                return

        # --- Table detection ---
        # First row looks like a header: distinct string values, not all numeric
        non_empty_headers = [v for v in first_row_strs if v]
        if non_empty_headers:
            all_numeric = all(v.replace(".", "").replace("-", "").isdigit() for v in non_empty_headers)
            distinct_headers = len(set(non_empty_headers)) == len(non_empty_headers)
            if not all_numeric and distinct_headers:
                # Check column consistency in data rows
                consistent_rows = 0
                for r in rows[1:]:
                    populated = sum(1 for v in r if v is not None and str(v).strip())
                    if populated > len(first_row) * 0.5:
                        consistent_rows += 1
                if consistent_rows > len(rows[1:]) * 0.5:
                    section.sub_structure = "table"
                    section.header_row = [str(v) if v is not None else "" for v in first_row]
                    return

        # Default: free_text (fail-closed)
        section.sub_structure = "free_text"

    # ------------------------------------------------------------------
    # Serialization
    # ------------------------------------------------------------------

    @staticmethod
    def _serialize_section(section: Section) -> str:
        """Dispatch to the appropriate sub-structure serializer."""
        if section.sub_structure == "table":
            return TextSerializer._serialize_table(section)
        if section.sub_structure == "checklist":
            return TextSerializer._serialize_checklist(section)
        if section.sub_structure == "matrix":
            return TextSerializer._serialize_matrix(section)
        return TextSerializer._serialize_free_text(section)

    @staticmethod
    def _serialize_table(section: Section) -> str:
        """Serialize a small table section as natural language sentences.

        Format: "In section '{title}', {col} is {val}, {col} is {val}."
        """
        headers = section.header_row or [
            f"Column {i}" for i in range(section.col_count)
        ]
        lines: list[str] = []
        for row in section.rows[1:]:  # skip header row
            parts = []
            for i, val in enumerate(row):
                col_name = headers[i] if i < len(headers) else f"Column {i}"
                display_val = "N/A" if val is None else str(val)
                parts.append(f"{col_name} is {display_val}")
            lines.append(f"In section '{section.title}', {', '.join(parts)}.")
        return "\n".join(lines)

    @staticmethod
    def _serialize_checklist(section: Section) -> str:
        """Serialize a checklist section.

        Format: "Item X: status is Y, due date is Z, responsible party is W."
        """
        headers = section.header_row or []
        headers_lower = [h.lower() for h in headers]

        # Identify column roles
        item_col: int | None = None
        status_col: int | None = None
        date_col: int | None = None
        responsible_col: int | None = None

        for i, h in enumerate(headers_lower):
            if h in _STATUS_KEYWORDS or "status" in h:
                status_col = i
            elif "date" in h or "due" in h:
                date_col = i
            elif "responsible" in h or "owner" in h or "assigned" in h:
                responsible_col = i
            elif item_col is None:
                item_col = i

        if item_col is None:
            item_col = 0

        lines: list[str] = []
        for row in section.rows[1:]:  # skip header row
            item_val = row[item_col] if item_col < len(row) else None
            item_str = str(item_val) if item_val is not None else "N/A"

            parts = [item_str]
            if status_col is not None and status_col < len(row):
                val = row[status_col]
                parts.append(f"status is {val if val is not None else 'N/A'}")
            if date_col is not None and date_col < len(row):
                val = row[date_col]
                parts.append(f"due date is {val if val is not None else 'N/A'}")
            if responsible_col is not None and responsible_col < len(row):
                val = row[responsible_col]
                parts.append(f"responsible party is {val if val is not None else 'N/A'}")

            lines.append(f"{parts[0]}: {', '.join(parts[1:])}." if len(parts) > 1 else f"{parts[0]}.")
        return "\n".join(lines)

    @staticmethod
    def _serialize_matrix(section: Section) -> str:
        """Serialize a matrix section.

        Format: "For {row_header}, {col_header} is {value}."
        """
        col_headers = [
            str(v) if v is not None else f"Column {i}"
            for i, v in enumerate(section.rows[0][1:], start=1)
        ]
        lines: list[str] = []
        for row in section.rows[1:]:
            row_header = str(row[0]) if row[0] is not None else "N/A"
            for j, val in enumerate(row[1:]):
                if val is not None and str(val).strip():
                    col_header = col_headers[j] if j < len(col_headers) else f"Column {j+1}"
                    lines.append(f"For {row_header}, {col_header} is {val}.")
        return "\n".join(lines)

    @staticmethod
    def _serialize_free_text(section: Section) -> str:
        """Serialize a free text section, preserving paragraph breaks."""
        paragraphs: list[str] = []
        for row in section.rows:
            cells = [str(v) for v in row if v is not None and str(v).strip()]
            if cells:
                paragraphs.append(" ".join(cells))
        content = "\n\n".join(paragraphs)
        if section.title:
            return f"{section.title}\n\n{content}"
        return content

    # ------------------------------------------------------------------
    # Error classification
    # ------------------------------------------------------------------

    @staticmethod
    def _classify_backend_error(exc: Exception) -> ErrorCode:
        """Map an exception to the most appropriate ErrorCode."""
        msg = str(exc).lower()
        if "timeout" in msg or "timed out" in msg:
            if "embed" in msg:
                return ErrorCode.E_BACKEND_EMBED_TIMEOUT
            if "vector" in msg or "qdrant" in msg or "collection" in msg:
                return ErrorCode.E_BACKEND_VECTOR_TIMEOUT
            return ErrorCode.E_PROCESS_SERIALIZE
        if "connect" in msg or "connection" in msg:
            if "embed" in msg:
                return ErrorCode.E_BACKEND_EMBED_CONNECT
            if "vector" in msg or "qdrant" in msg or "collection" in msg:
                return ErrorCode.E_BACKEND_VECTOR_CONNECT
            return ErrorCode.E_PROCESS_SERIALIZE
        # Default to serialize error for unknown exceptions
        return ErrorCode.E_PROCESS_SERIALIZE

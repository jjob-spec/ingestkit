"""Parser fallback chain for Excel file ingestion.

Implements a three-tier parsing strategy for ``.xlsx`` files:

1. **openpyxl** (full fidelity) -- merged cells, formulas, formatting, hidden status.
2. **pandas** ``read_excel`` (reduced fidelity) -- loses merged cell info, formulas, hidden status.
3. **openpyxl** ``data_only=True`` (minimal fidelity) -- cached values only, no formula info.

Fallback is per-sheet: each sheet independently traverses the chain, and the
:class:`~ingestkit_excel.models.ParserUsed` enum is recorded on each
:class:`~ingestkit_excel.models.SheetProfile`.  Non-fatal warnings
(``W_PARSER_FALLBACK``) and structured :class:`~ingestkit_excel.errors.IngestError`
objects are collected and returned alongside the
:class:`~ingestkit_excel.models.FileProfile`.
"""

from __future__ import annotations

import hashlib
import logging
import time
from pathlib import Path

import openpyxl
from openpyxl.chartsheet import Chartsheet
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

from ingestkit_excel.config import ExcelProcessorConfig
from ingestkit_excel.errors import ErrorCode, IngestError
from ingestkit_excel.models import FileProfile, ParserUsed, SheetProfile

logger = logging.getLogger("ingestkit_excel")


class ParserChain:
    """Three-tier fallback parser for ``.xlsx`` files.

    Tries openpyxl (full fidelity), then pandas ``read_excel`` (reduced
    fidelity), and finally openpyxl ``data_only=True`` (minimal fidelity)
    on a per-sheet basis.  Collects structured errors and warnings for
    every fallback or edge-case event.

    Parameters
    ----------
    config:
        Pipeline configuration controlling limits such as
        ``max_rows_in_memory`` and ``max_sample_rows``.
    """

    def __init__(self, config: ExcelProcessorConfig) -> None:
        self._config = config

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def parse(self, file_path: str) -> tuple[FileProfile, list[IngestError]]:
        """Parse an Excel file using the three-tier fallback chain.

        Parameters
        ----------
        file_path:
            Filesystem path to the ``.xlsx`` file.

        Returns
        -------
        tuple[FileProfile, list[IngestError]]
            A ``FileProfile`` aggregating all successfully parsed sheets,
            and a list of structured errors/warnings encountered during
            parsing.
        """
        errors: list[IngestError] = []
        sheet_profiles: list[SheetProfile] = []
        has_password = False
        has_chart_only = False

        start = time.monotonic()

        path = Path(file_path)

        # Verify file exists
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Compute content hash from raw bytes
        content_hash = self._compute_content_hash(file_path)

        # Check for zero-byte file
        file_size = path.stat().st_size
        if file_size == 0:
            errors.append(
                IngestError(
                    code=ErrorCode.E_PARSE_EMPTY,
                    message="File is empty (0 bytes).",
                    stage="parse",
                    recoverable=False,
                )
            )
            duration = time.monotonic() - start
            logger.error(
                "Parse failed: empty file %s (%.3fs)", file_path, duration
            )
            return (
                self._build_empty_file_profile(
                    file_path, file_size, content_hash
                ),
                errors,
            )

        # Attempt to open the workbook with openpyxl
        wb = None
        file_level_open_failed = False
        try:
            wb = openpyxl.load_workbook(file_path)
        except Exception as exc:
            exc_msg = str(exc).lower()
            if "password" in exc_msg or "encrypted" in exc_msg:
                has_password = True
                errors.append(
                    IngestError(
                        code=ErrorCode.E_PARSE_PASSWORD,
                        message=f"File is password-protected: {exc}",
                        stage="parse",
                        recoverable=False,
                    )
                )
                logger.error(
                    "Parse failed: password-protected file %s", file_path
                )
                return (
                    self._build_empty_file_profile(
                        file_path,
                        file_size,
                        content_hash,
                        has_password_protected_sheets=True,
                    ),
                    errors,
                )
            file_level_open_failed = True
            logger.warning(
                "openpyxl could not open file %s: %s", file_path, exc
            )

        if file_level_open_failed:
            # Try pandas to enumerate and parse sheets
            sheet_profiles, errors, has_chart_only = (
                self._parse_all_via_pandas_fallback(
                    file_path, errors
                )
            )
            if not sheet_profiles:
                # Try raw text fallback for the entire file
                sheet_profiles, errors = self._parse_all_via_raw_text_fallback(
                    file_path, errors
                )
            if not sheet_profiles:
                errors.append(
                    IngestError(
                        code=ErrorCode.E_PARSE_CORRUPT,
                        message="All parsers failed. File may be corrupt.",
                        stage="parse",
                        recoverable=False,
                    )
                )
                logger.error("Parse failed: corrupt file %s", file_path)
                return (
                    self._build_empty_file_profile(
                        file_path, file_size, content_hash
                    ),
                    errors,
                )
        else:
            # Successfully opened with openpyxl -- enumerate sheets
            assert wb is not None
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Check for chart-only sheets
                if isinstance(ws, Chartsheet):
                    has_chart_only = True
                    errors.append(
                        IngestError(
                            code=ErrorCode.W_SHEET_SKIPPED_CHART,
                            message=f"Sheet '{sheet_name}' is chart-only; skipped.",
                            sheet_name=sheet_name,
                            stage="parse",
                            recoverable=True,
                        )
                    )
                    logger.info(
                        "Skipped chart-only sheet '%s' in %s",
                        sheet_name,
                        file_path,
                    )
                    continue

                # Must be a Worksheet to proceed
                if not isinstance(ws, Worksheet):
                    continue

                # Check row count against max_rows_in_memory
                max_row = ws.max_row or 0
                if max_row > self._config.max_rows_in_memory:
                    errors.append(
                        IngestError(
                            code=ErrorCode.W_ROWS_TRUNCATED,
                            message=(
                                f"Sheet '{sheet_name}' has {max_row} rows, "
                                f"exceeding max_rows_in_memory "
                                f"({self._config.max_rows_in_memory}). "
                                f"Sheet skipped."
                            ),
                            sheet_name=sheet_name,
                            stage="parse",
                            recoverable=True,
                        )
                    )
                    logger.warning(
                        "Sheet '%s' exceeds max_rows_in_memory (%d > %d); skipped",
                        sheet_name,
                        max_row,
                        self._config.max_rows_in_memory,
                    )
                    continue

                # Tier 1: openpyxl primary
                profile = self._try_parse_sheet_openpyxl(ws, sheet_name)
                if profile is not None:
                    sheet_profiles.append(profile)
                    continue

                # Tier 1 failed -- record and try Tier 2: pandas fallback
                fallback_reason = ErrorCode.E_PARSE_OPENPYXL_FAIL.value
                logger.warning(
                    "openpyxl failed for sheet '%s'; trying pandas fallback",
                    sheet_name,
                )

                profile = self._try_parse_sheet_pandas(file_path, sheet_name)
                if profile is not None:
                    errors.append(
                        IngestError(
                            code=ErrorCode.W_PARSER_FALLBACK,
                            message=(
                                f"Sheet '{sheet_name}' parsed via pandas fallback. "
                                f"Reason: {fallback_reason}"
                            ),
                            sheet_name=sheet_name,
                            stage="parse",
                            recoverable=True,
                        )
                    )
                    sheet_profiles.append(profile)
                    continue

                # Tier 2 failed -- try Tier 3: raw text (data_only)
                fallback_reason = ErrorCode.E_PARSE_PANDAS_FAIL.value
                logger.warning(
                    "pandas failed for sheet '%s'; trying raw text fallback",
                    sheet_name,
                )

                profile = self._try_parse_sheet_raw_text(file_path, sheet_name)
                if profile is not None:
                    errors.append(
                        IngestError(
                            code=ErrorCode.W_PARSER_FALLBACK,
                            message=(
                                f"Sheet '{sheet_name}' parsed via raw text fallback. "
                                f"Reason: {fallback_reason}"
                            ),
                            sheet_name=sheet_name,
                            stage="parse",
                            recoverable=True,
                        )
                    )
                    sheet_profiles.append(profile)
                    continue

                # All three tiers failed for this sheet
                errors.append(
                    IngestError(
                        code=ErrorCode.E_PARSE_CORRUPT,
                        message=f"All parsers failed for sheet '{sheet_name}'.",
                        sheet_name=sheet_name,
                        stage="parse",
                        recoverable=False,
                    )
                )
                logger.error(
                    "All parsers failed for sheet '%s' in %s",
                    sheet_name,
                    file_path,
                )

            if wb is not None:
                wb.close()

        # If no data sheets were parsed, report empty
        if not sheet_profiles:
            # Only add E_PARSE_EMPTY if we haven't already added E_PARSE_CORRUPT
            has_corrupt = any(
                e.code == ErrorCode.E_PARSE_CORRUPT for e in errors
            )
            if not has_corrupt:
                errors.append(
                    IngestError(
                        code=ErrorCode.E_PARSE_EMPTY,
                        message="No data sheets found in workbook.",
                        stage="parse",
                        recoverable=False,
                    )
                )
            duration = time.monotonic() - start
            logger.info(
                "Parse of %s completed (no data sheets) in %.3fs",
                file_path,
                duration,
            )
            return (
                self._build_empty_file_profile(
                    file_path,
                    file_size,
                    content_hash,
                    has_password_protected_sheets=has_password,
                    has_chart_only_sheets=has_chart_only,
                ),
                errors,
            )

        # Build FileProfile
        file_profile = FileProfile(
            file_path=file_path,
            file_size_bytes=file_size,
            sheet_count=len(sheet_profiles),
            sheet_names=[sp.name for sp in sheet_profiles],
            sheets=sheet_profiles,
            has_password_protected_sheets=has_password,
            has_chart_only_sheets=has_chart_only,
            total_merged_cells=sum(sp.merged_cell_count for sp in sheet_profiles),
            total_rows=sum(sp.row_count for sp in sheet_profiles),
            content_hash=content_hash,
        )

        duration = time.monotonic() - start
        logger.info(
            "Parsed %s: %d sheets in %.3fs",
            file_path,
            len(sheet_profiles),
            duration,
        )

        return file_profile, errors

    # ------------------------------------------------------------------
    # Content hash
    # ------------------------------------------------------------------

    def _compute_content_hash(self, file_path: str) -> str:
        """Compute SHA-256 hex digest of the raw file bytes."""
        file_bytes = Path(file_path).read_bytes()
        return hashlib.sha256(file_bytes).hexdigest()

    # ------------------------------------------------------------------
    # Tier 1: openpyxl primary
    # ------------------------------------------------------------------

    def _try_parse_sheet_openpyxl(
        self, ws: Worksheet, sheet_name: str
    ) -> SheetProfile | None:
        """Attempt to parse a worksheet using openpyxl (full fidelity).

        Returns a ``SheetProfile`` on success, or ``None`` on failure.
        """
        try:
            return self._build_sheet_profile_from_openpyxl(ws, sheet_name)
        except Exception:
            logger.debug(
                "openpyxl parse failed for sheet '%s'", sheet_name, exc_info=True
            )
            return None

    def _build_sheet_profile_from_openpyxl(
        self, ws: Worksheet, sheet_name: str
    ) -> SheetProfile:
        """Extract all SheetProfile fields from an openpyxl Worksheet."""
        row_count = ws.max_row or 0
        col_count = ws.max_column or 0
        total_cells = row_count * col_count

        # Merged cells
        merged_cell_count = len(ws.merged_cells.ranges)
        merged_cell_ratio = (
            merged_cell_count / total_cells if total_cells > 0 else 0.0
        )

        # Read all cell data for analysis
        rows_data: list[list[object]] = []
        for row in ws.iter_rows(min_row=1, max_row=row_count, max_col=col_count):
            rows_data.append([cell.value for cell in row])

        # Formula detection
        has_formulas = False
        for row in ws.iter_rows(min_row=1, max_row=row_count, max_col=col_count):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    has_formulas = True
                    break
            if has_formulas:
                break

        # Hidden status
        is_hidden = ws.sheet_state == "hidden"

        # Trim to effective data bounds (removes inflated empty rows/cols)
        trimmed, first_data_row, _first_data_col = (
            self._trim_to_data_bounds(rows_data)
        )

        # Header detection on trimmed data
        header_row_detected, header_values, trimmed_header_idx = (
            self._detect_header(trimmed)
        )
        # Map index back to original rows_data coordinates
        header_row_index = (
            (first_data_row + trimmed_header_idx)
            if trimmed_header_idx is not None
            else None
        )

        # Compute ratios from DATA rows only (after header) so title/blank
        # rows don't pollute the signals that the inspector relies on.
        data_rows = (
            trimmed[trimmed_header_idx + 1:]
            if trimmed_header_idx is not None
            else trimmed
        )
        numeric_ratio, text_ratio, empty_ratio = self._compute_type_ratios(
            data_rows
        )
        column_type_consistency = self._compute_column_type_consistency(
            data_rows
        )

        # Sample rows (from trimmed data for relevance)
        sample_rows = self._extract_sample_rows(trimmed)

        return SheetProfile(
            name=sheet_name,
            row_count=row_count,
            col_count=col_count,
            merged_cell_count=merged_cell_count,
            merged_cell_ratio=merged_cell_ratio,
            header_row_detected=header_row_detected,
            header_row_index=header_row_index,
            header_values=header_values,
            column_type_consistency=column_type_consistency,
            numeric_ratio=numeric_ratio,
            text_ratio=text_ratio,
            empty_ratio=empty_ratio,
            sample_rows=sample_rows,
            has_formulas=has_formulas,
            is_hidden=is_hidden,
            parser_used=ParserUsed.OPENPYXL,
        )

    # ------------------------------------------------------------------
    # Tier 2: pandas fallback
    # ------------------------------------------------------------------

    def _try_parse_sheet_pandas(
        self, file_path: str, sheet_name: str
    ) -> SheetProfile | None:
        """Attempt to parse a sheet using pandas (reduced fidelity).

        Returns a ``SheetProfile`` on success, or ``None`` on failure.
        """
        try:
            return self._build_sheet_profile_from_pandas(file_path, sheet_name)
        except Exception:
            logger.debug(
                "pandas parse failed for sheet '%s'", sheet_name, exc_info=True
            )
            return None

    def _build_sheet_profile_from_pandas(
        self, file_path: str, sheet_name: str
    ) -> SheetProfile:
        """Extract SheetProfile fields from a pandas DataFrame."""
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        row_count = len(df)
        col_count = len(df.columns)

        # Convert to list-of-lists for shared analysis helpers
        rows_data: list[list[object]] = []
        for _, row in df.iterrows():
            rows_data.append([v if pd.notna(v) else None for v in row])

        trimmed, first_data_row, _first_data_col = (
            self._trim_to_data_bounds(rows_data)
        )
        header_row_detected, header_values, trimmed_header_idx = (
            self._detect_header(trimmed)
        )
        header_row_index = (
            (first_data_row + trimmed_header_idx)
            if trimmed_header_idx is not None
            else None
        )
        data_rows = (
            trimmed[trimmed_header_idx + 1:]
            if trimmed_header_idx is not None
            else trimmed
        )
        numeric_ratio, text_ratio, empty_ratio = self._compute_type_ratios(
            data_rows
        )
        column_type_consistency = self._compute_column_type_consistency(
            data_rows
        )
        sample_rows = self._extract_sample_rows(trimmed)

        return SheetProfile(
            name=sheet_name,
            row_count=row_count,
            col_count=col_count,
            merged_cell_count=0,
            merged_cell_ratio=0.0,
            header_row_detected=header_row_detected,
            header_row_index=header_row_index,
            header_values=header_values,
            column_type_consistency=column_type_consistency,
            numeric_ratio=numeric_ratio,
            text_ratio=text_ratio,
            empty_ratio=empty_ratio,
            sample_rows=sample_rows,
            has_formulas=False,
            is_hidden=False,
            parser_used=ParserUsed.PANDAS_FALLBACK,
        )

    # ------------------------------------------------------------------
    # Tier 3: openpyxl data_only (raw text fallback)
    # ------------------------------------------------------------------

    def _try_parse_sheet_raw_text(
        self, file_path: str, sheet_name: str
    ) -> SheetProfile | None:
        """Attempt to parse a sheet using openpyxl data_only mode (minimal fidelity).

        Returns a ``SheetProfile`` on success, or ``None`` on failure.
        """
        try:
            return self._build_sheet_profile_from_raw_text(
                file_path, sheet_name
            )
        except Exception:
            logger.debug(
                "raw text parse failed for sheet '%s'",
                sheet_name,
                exc_info=True,
            )
            return None

    def _build_sheet_profile_from_raw_text(
        self, file_path: str, sheet_name: str
    ) -> SheetProfile:
        """Extract SheetProfile fields using openpyxl with data_only=True."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        try:
            ws = wb[sheet_name]
            if not isinstance(ws, Worksheet):
                raise ValueError(
                    f"Sheet '{sheet_name}' is not a data worksheet."
                )

            row_count = ws.max_row or 0
            col_count = ws.max_column or 0
            total_cells = row_count * col_count

            merged_cell_count = len(ws.merged_cells.ranges)
            merged_cell_ratio = (
                merged_cell_count / total_cells if total_cells > 0 else 0.0
            )

            is_hidden = ws.sheet_state == "hidden"

            rows_data: list[list[object]] = []
            for row in ws.iter_rows(
                min_row=1, max_row=row_count, max_col=col_count
            ):
                rows_data.append([cell.value for cell in row])

            trimmed, first_data_row, _first_data_col = (
                self._trim_to_data_bounds(rows_data)
            )
            header_row_detected, header_values, trimmed_header_idx = (
                self._detect_header(trimmed)
            )
            header_row_index = (
                (first_data_row + trimmed_header_idx)
                if trimmed_header_idx is not None
                else None
            )
            data_rows = (
                trimmed[trimmed_header_idx + 1:]
                if trimmed_header_idx is not None
                else trimmed
            )
            numeric_ratio, text_ratio, empty_ratio = (
                self._compute_type_ratios(data_rows)
            )
            column_type_consistency = (
                self._compute_column_type_consistency(data_rows)
            )
            sample_rows = self._extract_sample_rows(trimmed)

            return SheetProfile(
                name=sheet_name,
                row_count=row_count,
                col_count=col_count,
                merged_cell_count=merged_cell_count,
                merged_cell_ratio=merged_cell_ratio,
                header_row_detected=header_row_detected,
                header_row_index=header_row_index,
                header_values=header_values,
                column_type_consistency=column_type_consistency,
                numeric_ratio=numeric_ratio,
                text_ratio=text_ratio,
                empty_ratio=empty_ratio,
                sample_rows=sample_rows,
                has_formulas=False,
                is_hidden=is_hidden,
                parser_used=ParserUsed.RAW_TEXT_FALLBACK,
            )
        finally:
            wb.close()

    # ------------------------------------------------------------------
    # File-level fallback helpers (when openpyxl can't even open the file)
    # ------------------------------------------------------------------

    def _parse_all_via_pandas_fallback(
        self,
        file_path: str,
        errors: list[IngestError],
    ) -> tuple[list[SheetProfile], list[IngestError], bool]:
        """Try to parse all sheets via pandas when openpyxl file-level open fails.

        Returns (sheet_profiles, errors, has_chart_only).
        """
        profiles: list[SheetProfile] = []
        has_chart_only = False
        try:
            all_dfs = pd.read_excel(
                file_path, sheet_name=None, header=None
            )
            for sname in all_dfs:
                sheet_name = str(sname)
                df = all_dfs[sname]
                rows_data: list[list[object]] = []
                for _, row in df.iterrows():
                    rows_data.append(
                        [v if pd.notna(v) else None for v in row]
                    )

                row_count = len(df)
                if row_count > self._config.max_rows_in_memory:
                    errors.append(
                        IngestError(
                            code=ErrorCode.W_ROWS_TRUNCATED,
                            message=(
                                f"Sheet '{sheet_name}' has {row_count} rows, "
                                f"exceeding max_rows_in_memory "
                                f"({self._config.max_rows_in_memory}). "
                                f"Sheet skipped."
                            ),
                            sheet_name=sheet_name,
                            stage="parse",
                            recoverable=True,
                        )
                    )
                    continue

                trimmed, first_data_row, _first_data_col = (
                    self._trim_to_data_bounds(rows_data)
                )
                header_row_detected, header_values, trimmed_header_idx = (
                    self._detect_header(trimmed)
                )
                header_row_index = (
                    (first_data_row + trimmed_header_idx)
                    if trimmed_header_idx is not None
                    else None
                )
                data_rows = (
                    trimmed[trimmed_header_idx + 1:]
                    if trimmed_header_idx is not None
                    else trimmed
                )
                numeric_ratio, text_ratio, empty_ratio = (
                    self._compute_type_ratios(data_rows)
                )
                column_type_consistency = (
                    self._compute_column_type_consistency(data_rows)
                )
                sample_rows = self._extract_sample_rows(trimmed)

                errors.append(
                    IngestError(
                        code=ErrorCode.W_PARSER_FALLBACK,
                        message=(
                            f"Sheet '{sheet_name}' parsed via pandas fallback. "
                            f"Reason: {ErrorCode.E_PARSE_OPENPYXL_FAIL.value}"
                        ),
                        sheet_name=sheet_name,
                        stage="parse",
                        recoverable=True,
                    )
                )

                profiles.append(
                    SheetProfile(
                        name=sheet_name,
                        row_count=row_count,
                        col_count=len(df.columns),
                        merged_cell_count=0,
                        merged_cell_ratio=0.0,
                        header_row_detected=header_row_detected,
                        header_row_index=header_row_index,
                        header_values=header_values,
                        column_type_consistency=column_type_consistency,
                        numeric_ratio=numeric_ratio,
                        text_ratio=text_ratio,
                        empty_ratio=empty_ratio,
                        sample_rows=sample_rows,
                        has_formulas=False,
                        is_hidden=False,
                        parser_used=ParserUsed.PANDAS_FALLBACK,
                    )
                )
        except Exception:
            logger.debug(
                "pandas file-level fallback failed for %s",
                file_path,
                exc_info=True,
            )

        return profiles, errors, has_chart_only

    def _parse_all_via_raw_text_fallback(
        self,
        file_path: str,
        errors: list[IngestError],
    ) -> tuple[list[SheetProfile], list[IngestError]]:
        """Try to parse all sheets via openpyxl data_only when pandas also fails."""
        profiles: list[SheetProfile] = []
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            try:
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if isinstance(ws, Chartsheet):
                        continue
                    if not isinstance(ws, Worksheet):
                        continue

                    max_row = ws.max_row or 0
                    if max_row > self._config.max_rows_in_memory:
                        errors.append(
                            IngestError(
                                code=ErrorCode.W_ROWS_TRUNCATED,
                                message=(
                                    f"Sheet '{sheet_name}' has {max_row} rows, "
                                    f"exceeding max_rows_in_memory."
                                ),
                                sheet_name=sheet_name,
                                stage="parse",
                                recoverable=True,
                            )
                        )
                        continue

                    profile = self._build_sheet_profile_from_raw_text(
                        file_path, sheet_name
                    )
                    errors.append(
                        IngestError(
                            code=ErrorCode.W_PARSER_FALLBACK,
                            message=(
                                f"Sheet '{sheet_name}' parsed via raw text fallback. "
                                f"Reason: {ErrorCode.E_PARSE_PANDAS_FAIL.value}"
                            ),
                            sheet_name=sheet_name,
                            stage="parse",
                            recoverable=True,
                        )
                    )
                    profiles.append(profile)
            finally:
                wb.close()
        except Exception:
            logger.debug(
                "raw text file-level fallback failed for %s",
                file_path,
                exc_info=True,
            )

        return profiles, errors

    # ------------------------------------------------------------------
    # Shared analysis helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _is_empty_cell(val: object) -> bool:
        """Return True if a cell value is logically empty."""
        if val is None:
            return True
        if isinstance(val, str) and val.strip() == "":
            return True
        return False

    def _trim_to_data_bounds(
        self, rows_data: list[list[object]]
    ) -> tuple[list[list[object]], int, int]:
        """Trim rows_data to the bounding box of non-empty cells.

        Real-world Excel files often have inflated ``max_row`` /
        ``max_column`` due to formatting, accidental edits, or empty
        merged regions.  This trims leading and trailing empty rows and
        columns so downstream analysis operates on actual data only.

        Returns (trimmed_rows, first_data_row, first_data_col) where
        ``first_data_row`` and ``first_data_col`` are 0-based offsets
        into the original ``rows_data`` so callers can map indices back.
        """
        if not rows_data:
            return [], 0, 0

        raw_col_count = max(len(r) for r in rows_data)

        # Find bounding box of non-empty cells
        min_row = len(rows_data)
        max_row = -1
        min_col = raw_col_count
        max_col = -1

        for row_idx, row in enumerate(rows_data):
            for col_idx, val in enumerate(row):
                if not self._is_empty_cell(val):
                    min_row = min(min_row, row_idx)
                    max_row = max(max_row, row_idx)
                    min_col = min(min_col, col_idx)
                    max_col = max(max_col, col_idx)

        if max_row == -1:
            # All empty
            return [], 0, 0

        trimmed: list[list[object]] = []
        for row_idx in range(min_row, max_row + 1):
            row = rows_data[row_idx]
            trimmed.append(row[min_col : max_col + 1])

        return trimmed, min_row, min_col

    def _detect_header(
        self, rows_data: list[list[object]]
    ) -> tuple[bool, list[str], int | None]:
        """Detect the header row by scanning past title and blank rows.

        Heuristic: scan the first ``_MAX_HEADER_SCAN_ROWS`` rows and
        return the first row where:

        * At least ``_MIN_HEADER_COLS`` cells are non-empty, **and**
        * every non-empty cell is a string, **and**
        * the **span-based fill ratio** (non-empty count / span from
          first to last non-empty cell) is ``>= 0.5``.

        Using the row's own span instead of the total column count makes
        this robust to inflated ``max_column`` and multi-region sheets
        where tables sit in different parts of the grid.

        Returns (header_detected, header_values, header_row_index).
        ``header_row_index`` is 0-based into ``rows_data``, or ``None``
        if no header was found.
        """
        _MAX_HEADER_SCAN_ROWS = 20
        _MIN_HEADER_COLS = 2

        if not rows_data:
            return False, [], None

        scan_limit = min(len(rows_data), _MAX_HEADER_SCAN_ROWS)
        for idx in range(scan_limit):
            row = rows_data[idx]
            non_null = [v for v in row if not self._is_empty_cell(v)]

            if len(non_null) < _MIN_HEADER_COLS:
                continue

            if not all(isinstance(v, str) for v in non_null):
                continue

            # Span-based fill ratio: first to last non-empty cell
            indices = [
                i for i, v in enumerate(row)
                if not self._is_empty_cell(v)
            ]
            span = indices[-1] - indices[0] + 1
            fill_ratio = len(non_null) / span if span > 0 else 0.0

            if fill_ratio >= 0.5:
                return True, [str(v) for v in non_null], idx

        return False, [], None

    def _compute_type_ratios(
        self, rows_data: list[list[object]]
    ) -> tuple[float, float, float]:
        """Compute numeric, text, and empty ratios across all cells.

        Returns (numeric_ratio, text_ratio, empty_ratio).
        """
        total = 0
        numeric_count = 0
        text_count = 0
        empty_count = 0

        for row in rows_data:
            for val in row:
                total += 1
                if val is None:
                    empty_count += 1
                elif isinstance(val, (int, float)):
                    numeric_count += 1
                elif isinstance(val, str):
                    if val.strip() == "":
                        empty_count += 1
                    else:
                        text_count += 1
                else:
                    # datetime, bool, etc. -- count as text
                    text_count += 1

        if total == 0:
            return 0.0, 0.0, 0.0

        return (
            numeric_count / total,
            text_count / total,
            empty_count / total,
        )

    def _compute_column_type_consistency(
        self, rows_data: list[list[object]]
    ) -> float:
        """Compute column type consistency.

        For each column, count the dominant non-null type.
        Consistency = sum of dominant counts / total non-null cells.
        """
        if not rows_data:
            return 0.0

        col_count = max(len(row) for row in rows_data) if rows_data else 0
        if col_count == 0:
            return 0.0

        total_non_null = 0
        dominant_sum = 0

        for col_idx in range(col_count):
            type_counts: dict[str, int] = {}
            for row in rows_data:
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val is None:
                    continue
                if isinstance(val, (int, float)):
                    type_key = "numeric"
                elif isinstance(val, str):
                    if val.strip() == "":
                        continue
                    type_key = "text"
                else:
                    type_key = "other"
                type_counts[type_key] = type_counts.get(type_key, 0) + 1

            col_non_null = sum(type_counts.values())
            total_non_null += col_non_null
            if type_counts:
                dominant_sum += max(type_counts.values())

        if total_non_null == 0:
            return 0.0

        return dominant_sum / total_non_null

    def _extract_sample_rows(
        self, rows_data: list[list[object]]
    ) -> list[list[str]]:
        """Extract the first N rows as lists of strings.

        N is determined by ``config.max_sample_rows``.
        """
        n = self._config.max_sample_rows
        sample: list[list[str]] = []
        for row in rows_data[:n]:
            sample.append([str(v) if v is not None else "" for v in row])
        return sample

    # ------------------------------------------------------------------
    # FileProfile builders
    # ------------------------------------------------------------------

    def _build_empty_file_profile(
        self,
        file_path: str,
        file_size_bytes: int,
        content_hash: str,
        has_password_protected_sheets: bool = False,
        has_chart_only_sheets: bool = False,
    ) -> FileProfile:
        """Build a minimal FileProfile when no sheets were parsed."""
        return FileProfile(
            file_path=file_path,
            file_size_bytes=file_size_bytes,
            sheet_count=0,
            sheet_names=[],
            sheets=[],
            has_password_protected_sheets=has_password_protected_sheets,
            has_chart_only_sheets=has_chart_only_sheets,
            total_merged_cells=0,
            total_rows=0,
            content_hash=content_hash,
        )

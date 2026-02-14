"""Tests for the shared test infrastructure in conftest.py.

Validates Protocol conformance of mock backends, MockLLM simulation modes,
.xlsx fixture loadability, and fixture availability.
"""

from __future__ import annotations

import openpyxl
import pandas as pd
import pytest

from ingestkit_core.models import BaseChunkMetadata, ChunkPayload
from ingestkit_core.protocols import (
    EmbeddingBackend,
    LLMBackend,
    StructuredDBBackend,
    VectorStoreBackend,
)

from .conftest import MockEmbedding, MockLLM, MockStructuredDB, MockVectorStore


# ---------------------------------------------------------------------------
# Protocol Conformance
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestProtocolConformance:
    """Every mock backend must satisfy its corresponding runtime-checkable Protocol."""

    def test_mock_vector_store_is_vector_store_backend(self) -> None:
        assert isinstance(MockVectorStore(), VectorStoreBackend)

    def test_mock_structured_db_is_structured_db_backend(self) -> None:
        assert isinstance(MockStructuredDB(), StructuredDBBackend)

    def test_mock_llm_is_llm_backend(self) -> None:
        assert isinstance(MockLLM(), LLMBackend)

    def test_mock_embedding_is_embedding_backend(self) -> None:
        assert isinstance(MockEmbedding(), EmbeddingBackend)


# ---------------------------------------------------------------------------
# MockVectorStore Behaviour
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestMockVectorStore:
    def test_upsert_and_count(self, mock_vector_store: MockVectorStore) -> None:
        mock_vector_store.ensure_collection("col", 768)
        chunk = _make_chunk("c1")
        assert mock_vector_store.upsert_chunks("col", [chunk]) == 1
        assert len(mock_vector_store.collections["col"]) == 1

    def test_delete_by_ids(self, mock_vector_store: MockVectorStore) -> None:
        mock_vector_store.ensure_collection("col", 768)
        chunks = [_make_chunk(f"c{i}") for i in range(3)]
        mock_vector_store.upsert_chunks("col", chunks)
        deleted = mock_vector_store.delete_by_ids("col", ["c0", "c2"])
        assert deleted == 2
        assert len(mock_vector_store.collections["col"]) == 1

    def test_delete_from_missing_collection(self, mock_vector_store: MockVectorStore) -> None:
        assert mock_vector_store.delete_by_ids("nope", ["x"]) == 0

    def test_create_payload_index(self, mock_vector_store: MockVectorStore) -> None:
        mock_vector_store.create_payload_index("col", "tenant_id", "keyword")
        assert ("tenant_id", "keyword") in mock_vector_store.indexes["col"]


# ---------------------------------------------------------------------------
# MockStructuredDB Behaviour
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestMockStructuredDB:
    def test_create_and_query_table(self, mock_structured_db: MockStructuredDB) -> None:
        df = pd.DataFrame({"a": [1, 2], "b": ["x", "y"]})
        mock_structured_db.create_table_from_dataframe("t1", df)
        assert mock_structured_db.table_exists("t1")
        schema = mock_structured_db.get_table_schema("t1")
        assert "a" in schema
        assert "b" in schema

    def test_drop_table(self, mock_structured_db: MockStructuredDB) -> None:
        df = pd.DataFrame({"x": [1]})
        mock_structured_db.create_table_from_dataframe("t1", df)
        mock_structured_db.drop_table("t1")
        assert not mock_structured_db.table_exists("t1")

    def test_get_connection_uri(self, mock_structured_db: MockStructuredDB) -> None:
        assert mock_structured_db.get_connection_uri() == "mock://in-memory"

    def test_schema_missing_table(self, mock_structured_db: MockStructuredDB) -> None:
        assert mock_structured_db.get_table_schema("nope") == {}


# ---------------------------------------------------------------------------
# MockLLM Simulation Modes
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestMockLLM:
    def test_valid_classify_response(self, mock_llm: MockLLM) -> None:
        expected = {"file_type": "tabular_data", "confidence": 0.85, "reasoning": "test"}
        mock_llm.enqueue_classify(expected)
        result = mock_llm.classify("prompt", "model")
        assert result == expected
        assert len(mock_llm.classify_calls) == 1

    def test_malformed_json_classify(self, mock_llm: MockLLM) -> None:
        mock_llm.enqueue_classify("__MALFORMED_JSON__")
        result = mock_llm.classify("prompt", "model")
        # The result is a dict wrapping raw text — NOT valid classification
        assert "raw" in result

    def test_timeout_classify(self, mock_llm: MockLLM) -> None:
        mock_llm.enqueue_classify("__TIMEOUT__")
        with pytest.raises(TimeoutError, match="simulated timeout"):
            mock_llm.classify("prompt", "model")

    def test_schema_invalid_classify(self, mock_llm: MockLLM) -> None:
        """A dict missing required fields is schema-invalid."""
        mock_llm.enqueue_classify({"unexpected_key": True})
        result = mock_llm.classify("prompt", "model")
        # Returns successfully but lacks expected keys
        assert "file_type" not in result

    def test_empty_queue_raises(self, mock_llm: MockLLM) -> None:
        with pytest.raises(RuntimeError, match="no classify responses"):
            mock_llm.classify("prompt", "model")

    def test_valid_generate_response(self, mock_llm: MockLLM) -> None:
        mock_llm.enqueue_generate("hello world")
        result = mock_llm.generate("prompt", "model")
        assert result == "hello world"
        assert len(mock_llm.generate_calls) == 1

    def test_timeout_generate(self, mock_llm: MockLLM) -> None:
        mock_llm.enqueue_generate("__TIMEOUT__")
        with pytest.raises(TimeoutError, match="simulated timeout"):
            mock_llm.generate("prompt", "model")

    def test_multiple_responses_dequeue_in_order(self, mock_llm: MockLLM) -> None:
        mock_llm.enqueue_classify({"a": 1}, {"b": 2})
        assert mock_llm.classify("p", "m") == {"a": 1}
        assert mock_llm.classify("p", "m") == {"b": 2}


# ---------------------------------------------------------------------------
# MockEmbedding Behaviour
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestMockEmbedding:
    def test_dimension(self, mock_embedding: MockEmbedding) -> None:
        assert mock_embedding.dimension() == 768

    def test_embed_returns_zero_vectors(self, mock_embedding: MockEmbedding) -> None:
        vectors = mock_embedding.embed(["hello", "world"])
        assert len(vectors) == 2
        assert all(v == 0.0 for v in vectors[0])
        assert len(vectors[0]) == 768

    def test_custom_dimension(self) -> None:
        emb = MockEmbedding(dim=384)
        assert emb.dimension() == 384
        vecs = emb.embed(["x"])
        assert len(vecs[0]) == 384

    def test_embed_records_calls(self, mock_embedding: MockEmbedding) -> None:
        mock_embedding.embed(["a", "b"])
        mock_embedding.embed(["c"])
        assert len(mock_embedding.embed_calls) == 2


# ---------------------------------------------------------------------------
# .xlsx Fixture Loadability
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestXlsxFixtures:
    def test_type_a_simple_loadable(self, type_a_simple_xlsx) -> None:
        wb = openpyxl.load_workbook(type_a_simple_xlsx)
        ws = wb.active
        assert ws.max_row >= 21  # header + 20 data rows
        assert ws.max_column == 3
        assert ws.cell(1, 1).value == "ID"

    def test_type_b_checklist_loadable(self, type_b_checklist_xlsx) -> None:
        wb = openpyxl.load_workbook(type_b_checklist_xlsx)
        ws = wb.active
        assert ws.cell(1, 1).value == "Compliance Checklist"
        # Verify merged cells exist
        assert len(ws.merged_cells.ranges) >= 1

    def test_type_c_hybrid_loadable(self, type_c_hybrid_xlsx) -> None:
        wb = openpyxl.load_workbook(type_c_hybrid_xlsx)
        assert len(wb.sheetnames) == 2
        # Sheet 1 is tabular
        ws1 = wb["Sales"]
        assert ws1.cell(1, 1).value == "Date"
        # Sheet 2 has merges
        ws2 = wb["Notes"]
        assert len(ws2.merged_cells.ranges) >= 1

    def test_edge_empty_loadable(self, edge_empty_xlsx) -> None:
        wb = openpyxl.load_workbook(edge_empty_xlsx)
        ws = wb.active
        # No data — max_row/max_column should be minimal
        assert ws.max_row is None or ws.max_row <= 1

    def test_edge_chart_only_loadable(self, edge_chart_only_xlsx) -> None:
        wb = openpyxl.load_workbook(edge_chart_only_xlsx)
        ws = wb.active
        assert ws.title == "ChartData"
        assert ws.max_row >= 6

    def test_edge_large_row_count(self, edge_large_xlsx) -> None:
        """Verify the large fixture has > 100,000 rows (exceeds max_rows_in_memory)."""
        wb = openpyxl.load_workbook(edge_large_xlsx, read_only=True)
        ws = wb.active
        row_count = sum(1 for _ in ws.rows)
        wb.close()
        assert row_count >= 100_001


# ---------------------------------------------------------------------------
# Fixture Availability (smoke tests)
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestFixtureAvailability:
    """Verify all expected fixtures are injected correctly by pytest."""

    def test_sample_config_available(self, sample_config) -> None:
        assert sample_config is not None
        assert sample_config.parser_version == "ingestkit_excel:1.0.0"

    def test_sample_ingest_key_available(self, sample_ingest_key) -> None:
        assert sample_ingest_key is not None
        assert sample_ingest_key.tenant_id == "test_tenant"

    def test_test_config_available(self, test_config) -> None:
        assert test_config is not None
        assert test_config.tenant_id == "test_tenant"
        assert test_config.log_sample_data is True

    def test_mock_vector_store_available(self, mock_vector_store) -> None:
        assert mock_vector_store is not None

    def test_mock_structured_db_available(self, mock_structured_db) -> None:
        assert mock_structured_db is not None

    def test_mock_llm_available(self, mock_llm) -> None:
        assert mock_llm is not None

    def test_mock_embedding_available(self, mock_embedding) -> None:
        assert mock_embedding is not None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_chunk(chunk_id: str) -> ChunkPayload:
    """Create a minimal ``ChunkPayload`` for testing."""
    return ChunkPayload(
        id=chunk_id,
        text="test text",
        vector=[0.0] * 768,
        metadata=BaseChunkMetadata(
            source_uri="file:///tmp/test.xlsx",
            source_format="xlsx",
            ingestion_method="test",
            parser_version="ingestkit_excel:1.0.0",
            chunk_index=0,
            chunk_hash="abc123",
            ingest_key="key123",
        ),
    )

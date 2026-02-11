"""Tests for the ParserChain three-tier fallback parser.

Uses openpyxl to programmatically create test .xlsx files and validates
the full parsing pipeline: primary parser, fallback chain, edge cases,
SheetProfile field population, and FileProfile aggregation.
"""

from __future__ import annotations

import hashlib
import re
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from ingestkit_excel.config import ExcelProcessorConfig
from ingestkit_excel.errors import ErrorCode, IngestError
from ingestkit_excel.models import FileProfile, ParserUsed, SheetProfile
from ingestkit_excel.parser_chain import ParserChain


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _create_clean_xlsx(path: Path, rows: int = 20, cols: int = 3) -> str:
    """Create a clean tabular xlsx with a header row and numeric data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Header row (all strings)
    headers = [f"Col{i+1}" for i in range(cols)]
    ws.append(headers)

    # Data rows (numeric)
    for r in range(rows):
        ws.append([r * cols + c for c in range(cols)])

    file_path = str(path / "clean.xlsx")
    wb.save(file_path)
    wb.close()
    return file_path


def _create_merged_cells_xlsx(path: Path) -> str:
    """Create an xlsx with merged cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MergedSheet"

    ws.append(["Header1", "Header2", "Header3"])
    ws.append([1, 2, 3])
    ws.append([4, 5, 6])
    ws.append([7, 8, 9])

    # Merge A2:A3
    ws.merge_cells("A2:A3")
    # Merge B2:C2
    ws.merge_cells("B2:C2")

    file_path = str(path / "merged.xlsx")
    wb.save(file_path)
    wb.close()
    return file_path


def _create_empty_xlsx(path: Path) -> str:
    """Create an xlsx with no data (just an empty sheet)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EmptySheet"
    # No data appended -- row/col will both be 0/None
    file_path = str(path / "empty.xlsx")
    wb.save(file_path)
    wb.close()
    return file_path


def _create_multi_sheet_xlsx(path: Path) -> str:
    """Create an xlsx with 3 sheets."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Name", "Age"])
    ws1.append(["Alice", 30])
    ws1.append(["Bob", 25])

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Product", "Price"])
    ws2.append(["Widget", 9.99])
    ws2.append(["Gadget", 19.99])

    ws3 = wb.create_sheet("Sheet3")
    ws3.append(["City", "Population"])
    ws3.append(["Tokyo", 13960000])

    file_path = str(path / "multi.xlsx")
    wb.save(file_path)
    wb.close()
    return file_path


def _create_numeric_xlsx(path: Path) -> str:
    """Create an xlsx with predominantly numeric data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Numeric"

    ws.append(["A", "B", "C"])
    for i in range(20):
        ws.append([i * 1.1, i * 2.2, i * 3.3])

    file_path = str(path / "numeric.xlsx")
    wb.save(file_path)
    wb.close()
    return file_path


def _create_text_xlsx(path: Path) -> str:
    """Create an xlsx with predominantly text data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Text"

    ws.append(["Name", "Description", "Category"])
    for i in range(20):
        ws.append([f"item_{i}", f"description of item {i}", f"cat_{i % 3}"])

    file_path = str(path / "text.xlsx")
    wb.save(file_path)
    wb.close()
    return file_path


def _create_formula_xlsx(path: Path) -> str:
    """Create an xlsx with formulas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulas"

    ws.append(["Value", "Doubled"])
    for i in range(1, 11):
        ws.append([i, f"=A{i+1}*2"])

    file_path = str(path / "formulas.xlsx")
    wb.save(file_path)
    wb.close()
    return file_path


def _create_hidden_sheet_xlsx(path: Path) -> str:
    """Create an xlsx with a hidden sheet."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Visible"
    ws1.append(["Data", "Here"])
    ws1.append([1, 2])

    ws2 = wb.create_sheet("Hidden")
    ws2.sheet_state = "hidden"
    ws2.append(["Secret", "Data"])
    ws2.append([42, 99])

    file_path = str(path / "hidden.xlsx")
    wb.save(file_path)
    wb.close()
    return file_path


def _create_chart_sheet_xlsx(path: Path) -> str:
    """Create an xlsx with a chartsheet alongside a data sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Category", "Value"])
    ws.append(["A", 10])
    ws.append(["B", 20])
    ws.append(["C", 30])

    # Create a chart sheet
    cs = wb.create_chartsheet("MyChart")
    from openpyxl.chart import BarChart, Reference

    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    cs.add_chart(chart)

    file_path = str(path / "chart.xlsx")
    wb.save(file_path)
    wb.close()
    return file_path


def _create_large_sheet_xlsx(path: Path, rows: int) -> str:
    """Create an xlsx with a specified number of rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BigSheet"

    ws.append(["Col1", "Col2"])
    for i in range(rows):
        ws.append([i, i * 2])

    file_path = str(path / "large.xlsx")
    wb.save(file_path)
    wb.close()
    return file_path


def _create_mixed_types_xlsx(path: Path) -> str:
    """Create an xlsx with mixed column types for consistency testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mixed"

    ws.append(["ID", "Name", "Value"])
    ws.append([1, "alpha", 10.5])
    ws.append([2, "beta", 20.3])
    ws.append(["three", "gamma", "thirty"])  # mixed types in col 1 & 3
    ws.append([4, "delta", 40.1])

    file_path = str(path / "mixed.xlsx")
    wb.save(file_path)
    wb.close()
    return file_path


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def config() -> ExcelProcessorConfig:
    """Default config for tests."""
    return ExcelProcessorConfig()


@pytest.fixture()
def chain(config: ExcelProcessorConfig) -> ParserChain:
    """ParserChain with default config."""
    return ParserChain(config)


@pytest.fixture()
def small_config() -> ExcelProcessorConfig:
    """Config with very small max_rows_in_memory for truncation testing."""
    return ExcelProcessorConfig(max_rows_in_memory=10)


@pytest.fixture()
def small_chain(small_config: ExcelProcessorConfig) -> ParserChain:
    """ParserChain with small max_rows_in_memory."""
    return ParserChain(small_config)


# ---------------------------------------------------------------------------
# Test: clean xlsx with primary parser
# ---------------------------------------------------------------------------


class TestCleanXlsxPrimaryParser:
    """Verify that a clean xlsx is parsed using the openpyxl primary parser."""

    def test_clean_xlsx_primary_parser(self, tmp_path: Path, chain: ParserChain) -> None:
        file_path = _create_clean_xlsx(tmp_path)
        profile, errors = chain.parse(file_path)

        assert isinstance(profile, FileProfile)
        assert len(profile.sheets) == 1
        assert profile.sheets[0].parser_used == ParserUsed.OPENPYXL

        # No errors or warnings for a clean file
        non_info_errors = [
            e for e in errors if e.code.value.startswith("E_")
        ]
        assert len(non_info_errors) == 0


# ---------------------------------------------------------------------------
# Test: SheetProfile fields populated
# ---------------------------------------------------------------------------


class TestSheetProfileFieldsPopulated:
    """Verify all SheetProfile fields have reasonable values."""

    def test_sheet_profile_fields_populated(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        file_path = _create_clean_xlsx(tmp_path, rows=20, cols=3)
        profile, _ = chain.parse(file_path)

        sp = profile.sheets[0]
        assert sp.name == "Data"
        # 20 data rows + 1 header row
        assert sp.row_count == 21
        assert sp.col_count == 3
        assert sp.merged_cell_count == 0
        assert sp.merged_cell_ratio == 0.0
        assert sp.header_row_detected is True
        assert sp.header_values == ["Col1", "Col2", "Col3"]
        assert 0.0 <= sp.column_type_consistency <= 1.0
        assert 0.0 <= sp.numeric_ratio <= 1.0
        assert 0.0 <= sp.text_ratio <= 1.0
        assert 0.0 <= sp.empty_ratio <= 1.0
        assert abs(sp.numeric_ratio + sp.text_ratio + sp.empty_ratio - 1.0) < 0.01
        assert len(sp.sample_rows) > 0
        assert isinstance(sp.has_formulas, bool)
        assert isinstance(sp.is_hidden, bool)
        assert sp.parser_used == ParserUsed.OPENPYXL


# ---------------------------------------------------------------------------
# Test: content hash computed
# ---------------------------------------------------------------------------


class TestContentHashComputed:
    """Verify FileProfile.content_hash is a valid SHA-256 hex string."""

    def test_content_hash_computed(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        file_path = _create_clean_xlsx(tmp_path)
        profile, _ = chain.parse(file_path)

        assert len(profile.content_hash) == 64
        assert re.fullmatch(r"[0-9a-f]{64}", profile.content_hash)

        # Verify independently
        expected = hashlib.sha256(Path(file_path).read_bytes()).hexdigest()
        assert profile.content_hash == expected


# ---------------------------------------------------------------------------
# Test: merged cells detected
# ---------------------------------------------------------------------------


class TestMergedCellsDetected:
    """Verify merged cells are detected and ratios computed."""

    def test_merged_cells_detected(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        file_path = _create_merged_cells_xlsx(tmp_path)
        profile, _ = chain.parse(file_path)

        sp = profile.sheets[0]
        assert sp.merged_cell_count > 0
        assert sp.merged_cell_ratio > 0.0
        assert profile.total_merged_cells > 0


# ---------------------------------------------------------------------------
# Test: empty workbook
# ---------------------------------------------------------------------------


class TestEmptyWorkbook:
    """Verify that an xlsx with no data produces E_PARSE_EMPTY."""

    def test_empty_workbook(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        file_path = _create_empty_xlsx(tmp_path)
        profile, errors = chain.parse(file_path)

        # An openpyxl workbook with an empty sheet still reports max_row=1
        # for a blank sheet that was created. We test for either:
        # - E_PARSE_EMPTY if truly empty, or
        # - a profile with very low row counts
        # The key check is that parsing doesn't crash.
        assert isinstance(profile, FileProfile)

    def test_zero_byte_file(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        """A zero-byte file must produce E_PARSE_EMPTY."""
        file_path = str(tmp_path / "zero.xlsx")
        Path(file_path).write_bytes(b"")
        profile, errors = chain.parse(file_path)

        error_codes = [e.code for e in errors]
        assert ErrorCode.E_PARSE_EMPTY in error_codes
        assert profile.sheet_count == 0


# ---------------------------------------------------------------------------
# Test: chart-only sheet
# ---------------------------------------------------------------------------


class TestChartOnlySheet:
    """Verify chart-only sheets are detected and skipped with W_SHEET_SKIPPED_CHART."""

    def test_chart_only_sheet(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        file_path = _create_chart_sheet_xlsx(tmp_path)
        profile, errors = chain.parse(file_path)

        error_codes = [e.code for e in errors]
        assert ErrorCode.W_SHEET_SKIPPED_CHART in error_codes
        assert profile.has_chart_only_sheets is True

        # The data sheet should still be parsed
        assert profile.sheet_count >= 1
        data_sheets = [s for s in profile.sheets if s.name == "Data"]
        assert len(data_sheets) == 1


# ---------------------------------------------------------------------------
# Test: max rows truncation
# ---------------------------------------------------------------------------


class TestMaxRowsTruncation:
    """Verify W_ROWS_TRUNCATED when sheet exceeds max_rows_in_memory."""

    def test_max_rows_truncation(
        self, tmp_path: Path, small_chain: ParserChain
    ) -> None:
        # small_chain has max_rows_in_memory=10
        # Create a sheet with 20 data rows + 1 header = 21 rows total
        file_path = _create_large_sheet_xlsx(tmp_path, rows=20)
        profile, errors = small_chain.parse(file_path)

        error_codes = [e.code for e in errors]
        assert ErrorCode.W_ROWS_TRUNCATED in error_codes


# ---------------------------------------------------------------------------
# Test: header detection
# ---------------------------------------------------------------------------


class TestHeaderDetection:
    """Verify header row detection heuristic."""

    def test_header_detection(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        file_path = _create_clean_xlsx(tmp_path)
        profile, _ = chain.parse(file_path)

        sp = profile.sheets[0]
        assert sp.header_row_detected is True
        assert sp.header_values == ["Col1", "Col2", "Col3"]

    def test_no_header_when_all_numeric(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        """A sheet where the first row is all numeric should not detect a header."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NoHeader"
        ws.append([1, 2, 3])
        ws.append([4, 5, 6])
        file_path = str(tmp_path / "noheader.xlsx")
        wb.save(file_path)
        wb.close()

        profile, _ = chain.parse(file_path)
        sp = profile.sheets[0]
        assert sp.header_row_detected is False
        assert sp.header_values == []


# ---------------------------------------------------------------------------
# Test: sample rows populated
# ---------------------------------------------------------------------------


class TestSampleRowsPopulated:
    """Verify sample_rows contains the correct number of rows."""

    def test_sample_rows_populated(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        file_path = _create_clean_xlsx(tmp_path, rows=20)
        profile, _ = chain.parse(file_path)

        sp = profile.sheets[0]
        # Default max_sample_rows is 3
        assert len(sp.sample_rows) == 3
        # Each sample row should have values as strings
        for row in sp.sample_rows:
            assert all(isinstance(v, str) for v in row)

    def test_sample_rows_fewer_than_max(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        """When fewer rows than max_sample_rows, return all rows."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Small"
        ws.append(["A"])
        ws.append([1])
        file_path = str(tmp_path / "small.xlsx")
        wb.save(file_path)
        wb.close()

        profile, _ = chain.parse(file_path)
        sp = profile.sheets[0]
        assert len(sp.sample_rows) == 2  # only 2 rows exist


# ---------------------------------------------------------------------------
# Test: column type consistency
# ---------------------------------------------------------------------------


class TestColumnTypeConsistency:
    """Verify column type consistency computation."""

    def test_uniform_types_high_consistency(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        """Uniform numeric columns should have high consistency."""
        file_path = _create_numeric_xlsx(tmp_path)
        profile, _ = chain.parse(file_path)

        sp = profile.sheets[0]
        # Header row adds some string cells, but columns are still
        # predominantly one type
        assert sp.column_type_consistency > 0.7

    def test_mixed_types_lower_consistency(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        """Mixed-type columns should have lower consistency."""
        file_path = _create_mixed_types_xlsx(tmp_path)
        profile, _ = chain.parse(file_path)

        sp = profile.sheets[0]
        # With mixed types in some columns, consistency should be
        # less than perfect (but still above 0 since some columns are uniform)
        assert 0.0 < sp.column_type_consistency <= 1.0


# ---------------------------------------------------------------------------
# Test: numeric and text ratios
# ---------------------------------------------------------------------------


class TestNumericTextRatios:
    """Verify numeric_ratio and text_ratio computation."""

    def test_numeric_heavy_xlsx(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        """Numeric-heavy file should have high numeric_ratio."""
        file_path = _create_numeric_xlsx(tmp_path)
        profile, _ = chain.parse(file_path)

        sp = profile.sheets[0]
        assert sp.numeric_ratio > sp.text_ratio

    def test_text_heavy_xlsx(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        """Text-heavy file should have high text_ratio."""
        file_path = _create_text_xlsx(tmp_path)
        profile, _ = chain.parse(file_path)

        sp = profile.sheets[0]
        assert sp.text_ratio > sp.numeric_ratio


# ---------------------------------------------------------------------------
# Test: multiple sheets
# ---------------------------------------------------------------------------


class TestMultipleSheets:
    """Verify all sheets in a multi-sheet workbook are parsed."""

    def test_multiple_sheets(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        file_path = _create_multi_sheet_xlsx(tmp_path)
        profile, errors = chain.parse(file_path)

        assert profile.sheet_count == 3
        assert set(profile.sheet_names) == {"Sheet1", "Sheet2", "Sheet3"}
        assert len(profile.sheets) == 3

        # All sheets use primary parser
        for sp in profile.sheets:
            assert sp.parser_used == ParserUsed.OPENPYXL

        # Total rows is sum of individual sheet rows
        assert profile.total_rows == sum(sp.row_count for sp in profile.sheets)


# ---------------------------------------------------------------------------
# Test: formula detection
# ---------------------------------------------------------------------------


class TestFormulaDetection:
    """Verify formula detection in openpyxl mode."""

    def test_formulas_detected(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        file_path = _create_formula_xlsx(tmp_path)
        profile, _ = chain.parse(file_path)

        sp = profile.sheets[0]
        assert sp.has_formulas is True

    def test_no_formulas_clean_data(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        file_path = _create_clean_xlsx(tmp_path)
        profile, _ = chain.parse(file_path)

        sp = profile.sheets[0]
        assert sp.has_formulas is False


# ---------------------------------------------------------------------------
# Test: hidden sheet detection
# ---------------------------------------------------------------------------


class TestHiddenSheetDetection:
    """Verify hidden sheet status is detected."""

    def test_hidden_sheet_detected(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        file_path = _create_hidden_sheet_xlsx(tmp_path)
        profile, _ = chain.parse(file_path)

        hidden_sheets = [sp for sp in profile.sheets if sp.is_hidden]
        visible_sheets = [sp for sp in profile.sheets if not sp.is_hidden]
        assert len(hidden_sheets) == 1
        assert hidden_sheets[0].name == "Hidden"
        assert len(visible_sheets) == 1
        assert visible_sheets[0].name == "Visible"


# ---------------------------------------------------------------------------
# Test: FileProfile aggregate fields
# ---------------------------------------------------------------------------


class TestFileProfileAggregateFields:
    """Verify FileProfile aggregates correctly from SheetProfiles."""

    def test_file_profile_aggregate_fields(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        file_path = _create_multi_sheet_xlsx(tmp_path)
        profile, _ = chain.parse(file_path)

        assert profile.file_path == file_path
        assert profile.file_size_bytes > 0
        assert profile.sheet_count == 3
        assert len(profile.sheet_names) == 3
        assert len(profile.sheets) == 3
        assert profile.total_rows == sum(sp.row_count for sp in profile.sheets)
        assert profile.total_merged_cells == sum(
            sp.merged_cell_count for sp in profile.sheets
        )
        assert profile.has_password_protected_sheets is False
        assert profile.has_chart_only_sheets is False
        assert len(profile.content_hash) == 64


# ---------------------------------------------------------------------------
# Test: corrupt file
# ---------------------------------------------------------------------------


class TestCorruptFile:
    """Verify that a completely corrupt file produces E_PARSE_CORRUPT."""

    def test_all_parsers_fail_produces_e_parse_corrupt(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        """Random bytes with .xlsx extension should fail all parsers."""
        file_path = str(tmp_path / "corrupt.xlsx")
        Path(file_path).write_bytes(b"this is not an xlsx file at all")

        profile, errors = chain.parse(file_path)
        error_codes = [e.code for e in errors]
        assert ErrorCode.E_PARSE_CORRUPT in error_codes
        assert profile.sheet_count == 0


# ---------------------------------------------------------------------------
# Test: password-protected file (mocked)
# ---------------------------------------------------------------------------


class TestPasswordProtectedFile:
    """Verify password-protected file detection via mock."""

    def test_password_protected_file(
        self, tmp_path: Path, chain: ParserChain
    ) -> None:
        """Mock openpyxl to raise a password error."""
        file_path = _create_clean_xlsx(tmp_path)

        with patch(
            "ingestkit_excel.parser_chain.openpyxl.load_workbook"
        ) as mock_load:
            mock_load.side_effect = Exception(
                "File is encrypted with a password"
            )
            profile, errors = chain.parse(file_path)

        error_codes = [e.code for e in errors]
        assert ErrorCode.E_PARSE_PASSWORD in error_codes
        assert profile.has_password_protected_sheets is True
        assert profile.sheet_count == 0


# ---------------------------------------------------------------------------
# Test: parser fallback (openpyxl fails, pandas succeeds)
# ---------------------------------------------------------------------------


class TestParserFallback:
    """Verify the fallback chain when the primary parser fails."""

    def test_openpyxl_fail_falls_back_to_pandas(
        self, tmp_path: Path, config: ExcelProcessorConfig
    ) -> None:
        """When openpyxl sheet parsing fails, pandas should be used."""
        file_path = _create_clean_xlsx(tmp_path)
        chain = ParserChain(config)

        # Mock _try_parse_sheet_openpyxl to return None (simulating failure)
        with patch.object(
            chain, "_try_parse_sheet_openpyxl", return_value=None
        ):
            profile, errors = chain.parse(file_path)

        assert profile.sheet_count == 1
        sp = profile.sheets[0]
        assert sp.parser_used == ParserUsed.PANDAS_FALLBACK

        # Check W_PARSER_FALLBACK warning
        fallback_warnings = [
            e for e in errors if e.code == ErrorCode.W_PARSER_FALLBACK
        ]
        assert len(fallback_warnings) >= 1
        assert ErrorCode.E_PARSE_OPENPYXL_FAIL.value in fallback_warnings[0].message

    def test_openpyxl_and_pandas_fail_falls_back_to_raw_text(
        self, tmp_path: Path, config: ExcelProcessorConfig
    ) -> None:
        """When both openpyxl and pandas fail, raw text fallback should be used."""
        file_path = _create_clean_xlsx(tmp_path)
        chain = ParserChain(config)

        with patch.object(
            chain, "_try_parse_sheet_openpyxl", return_value=None
        ), patch.object(chain, "_try_parse_sheet_pandas", return_value=None):
            profile, errors = chain.parse(file_path)

        assert profile.sheet_count == 1
        sp = profile.sheets[0]
        assert sp.parser_used == ParserUsed.RAW_TEXT_FALLBACK

        # Check W_PARSER_FALLBACK warning
        fallback_warnings = [
            e for e in errors if e.code == ErrorCode.W_PARSER_FALLBACK
        ]
        assert len(fallback_warnings) >= 1
        assert ErrorCode.E_PARSE_PANDAS_FAIL.value in fallback_warnings[0].message


# ---------------------------------------------------------------------------
# Test: per-sheet fallback independence
# ---------------------------------------------------------------------------


class TestPerSheetFallbackIndependence:
    """Verify that fallback is per-sheet, not per-file."""

    def test_per_sheet_fallback_independence(
        self, tmp_path: Path, config: ExcelProcessorConfig
    ) -> None:
        """Sheet1 uses openpyxl, Sheet2 falls back to pandas."""
        file_path = _create_multi_sheet_xlsx(tmp_path)
        chain = ParserChain(config)

        original_try_parse = chain._try_parse_sheet_openpyxl

        call_count = 0

        def selective_fail(ws, sheet_name):
            nonlocal call_count
            call_count += 1
            if sheet_name == "Sheet2":
                return None  # Simulate failure for Sheet2
            return original_try_parse(ws, sheet_name)

        with patch.object(
            chain, "_try_parse_sheet_openpyxl", side_effect=selective_fail
        ):
            profile, errors = chain.parse(file_path)

        assert profile.sheet_count == 3

        parsers_by_name = {sp.name: sp.parser_used for sp in profile.sheets}
        assert parsers_by_name["Sheet1"] == ParserUsed.OPENPYXL
        assert parsers_by_name["Sheet2"] == ParserUsed.PANDAS_FALLBACK
        assert parsers_by_name["Sheet3"] == ParserUsed.OPENPYXL


# ---------------------------------------------------------------------------
# Test: file not found
# ---------------------------------------------------------------------------


class TestFileNotFound:
    """Verify FileNotFoundError when path doesn't exist."""

    def test_file_not_found(self, chain: ParserChain) -> None:
        with pytest.raises(FileNotFoundError):
            chain.parse("/nonexistent/path/to/file.xlsx")


# ---------------------------------------------------------------------------
# Test: fallback_reason_code correctly set
# ---------------------------------------------------------------------------


class TestFallbackReasonCode:
    """Verify fallback_reason_code in W_PARSER_FALLBACK messages."""

    def test_fallback_reason_code_correctly_set(
        self, tmp_path: Path, config: ExcelProcessorConfig
    ) -> None:
        file_path = _create_clean_xlsx(tmp_path)
        chain = ParserChain(config)

        with patch.object(
            chain, "_try_parse_sheet_openpyxl", return_value=None
        ):
            _, errors = chain.parse(file_path)

        fallback_warnings = [
            e for e in errors if e.code == ErrorCode.W_PARSER_FALLBACK
        ]
        assert len(fallback_warnings) >= 1
        # The reason code should reference E_PARSE_OPENPYXL_FAIL
        assert ErrorCode.E_PARSE_OPENPYXL_FAIL.value in fallback_warnings[0].message

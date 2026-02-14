"""Shared test fixtures for ingestkit-excel tests.

Provides mock backends that satisfy the four Protocol interfaces
(VectorStoreBackend, StructuredDBBackend, LLMBackend, EmbeddingBackend),
a ``test_config`` fixture, and session-scoped .xlsx file generators.
"""

from __future__ import annotations

import json
import pathlib
import tempfile
from typing import Any

import openpyxl
import pandas as pd
import pytest

from ingestkit_core.models import ChunkPayload
from ingestkit_excel.config import ExcelProcessorConfig
from ingestkit_excel.models import IngestKey


# ---------------------------------------------------------------------------
# Existing fixtures (preserved from original conftest.py)
# ---------------------------------------------------------------------------


@pytest.fixture()
def sample_config() -> ExcelProcessorConfig:
    """Return an ExcelProcessorConfig with all defaults."""
    return ExcelProcessorConfig()


@pytest.fixture()
def sample_ingest_key() -> IngestKey:
    """Return a sample IngestKey instance for testing."""
    return IngestKey(
        content_hash="abc123def456",
        source_uri="file:///tmp/test.xlsx",
        parser_version="ingestkit_excel:1.0.0",
        tenant_id="test_tenant",
    )


# ---------------------------------------------------------------------------
# Mock Backends
# ---------------------------------------------------------------------------


class MockVectorStore:
    """In-memory vector store satisfying ``VectorStoreBackend`` protocol.

    Stores chunks in a plain list, keyed by collection name.
    """

    def __init__(self) -> None:
        self.collections: dict[str, list[ChunkPayload]] = {}
        self.indexes: dict[str, list[tuple[str, str]]] = {}

    def upsert_chunks(self, collection: str, chunks: list[ChunkPayload]) -> int:
        self.collections.setdefault(collection, []).extend(chunks)
        return len(chunks)

    def ensure_collection(self, collection: str, vector_size: int) -> None:
        if collection not in self.collections:
            self.collections[collection] = []

    def create_payload_index(self, collection: str, field: str, field_type: str) -> None:
        self.indexes.setdefault(collection, []).append((field, field_type))

    def delete_by_ids(self, collection: str, ids: list[str]) -> int:
        if collection not in self.collections:
            return 0
        before = len(self.collections[collection])
        self.collections[collection] = [
            c for c in self.collections[collection] if c.id not in set(ids)
        ]
        return before - len(self.collections[collection])


class MockStructuredDB:
    """In-memory structured DB satisfying ``StructuredDBBackend`` protocol.

    Stores DataFrames in a dict keyed by table name.
    """

    def __init__(self) -> None:
        self.tables: dict[str, pd.DataFrame] = {}

    def create_table_from_dataframe(self, table_name: str, df: pd.DataFrame) -> None:
        self.tables[table_name] = df.copy()

    def drop_table(self, table_name: str) -> None:
        self.tables.pop(table_name, None)

    def table_exists(self, table_name: str) -> bool:
        return table_name in self.tables

    def get_table_schema(self, table_name: str) -> dict:
        if table_name not in self.tables:
            return {}
        return {col: str(dtype) for col, dtype in self.tables[table_name].dtypes.items()}

    def get_connection_uri(self) -> str:
        return "mock://in-memory"


class MockLLM:
    """Queue-based LLM backend satisfying ``LLMBackend`` protocol.

    Push responses (dicts for classify, strings for generate) onto
    ``classify_responses`` / ``generate_responses``.  Each call pops from
    the front of the queue.

    Special sentinel values:
    - ``"__MALFORMED_JSON__"`` in classify queue: returns a non-JSON string.
    - ``"__TIMEOUT__"`` in either queue: raises ``TimeoutError``.
    - Any dict missing required keys simulates schema-invalid output.
    """

    def __init__(self) -> None:
        self.classify_responses: list[Any] = []
        self.generate_responses: list[str] = []
        self.classify_calls: list[dict[str, Any]] = []
        self.generate_calls: list[dict[str, Any]] = []

    # -- helpers to enqueue responses --

    def enqueue_classify(self, *responses: Any) -> None:
        self.classify_responses.extend(responses)

    def enqueue_generate(self, *responses: str) -> None:
        self.generate_responses.extend(responses)

    # -- protocol methods --

    def classify(
        self,
        prompt: str,
        model: str,
        temperature: float = 0.1,
        timeout: float | None = None,
    ) -> dict:
        self.classify_calls.append(
            {"prompt": prompt, "model": model, "temperature": temperature, "timeout": timeout}
        )
        if not self.classify_responses:
            raise RuntimeError("MockLLM: no classify responses enqueued")
        response = self.classify_responses.pop(0)

        if response == "__TIMEOUT__":
            raise TimeoutError("MockLLM simulated timeout")
        if response == "__MALFORMED_JSON__":
            # Return a raw string that is NOT valid JSON — downstream
            # code calling json.loads() on it will get a ValueError.
            # The protocol says classify returns dict, so we return a
            # dict wrapping the raw text to stay protocol-compliant
            # while still triggering downstream parse errors.
            return {"raw": "<<<not json>>>"}
        return response

    def generate(
        self,
        prompt: str,
        model: str,
        temperature: float = 0.7,
        timeout: float | None = None,
    ) -> str:
        self.generate_calls.append(
            {"prompt": prompt, "model": model, "temperature": temperature, "timeout": timeout}
        )
        if not self.generate_responses:
            raise RuntimeError("MockLLM: no generate responses enqueued")
        response = self.generate_responses.pop(0)

        if response == "__TIMEOUT__":
            raise TimeoutError("MockLLM simulated timeout")
        return response


class MockEmbedding:
    """Zero-vector embedding backend satisfying ``EmbeddingBackend`` protocol.

    Returns ``[0.0] * dim`` for every input text.
    """

    def __init__(self, dim: int = 768) -> None:
        self._dim = dim
        self.embed_calls: list[list[str]] = []

    def embed(self, texts: list[str], timeout: float | None = None) -> list[list[float]]:
        self.embed_calls.append(texts)
        return [[0.0] * self._dim for _ in texts]

    def dimension(self) -> int:
        return self._dim


# ---------------------------------------------------------------------------
# Mock Backend Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def mock_vector_store() -> MockVectorStore:
    """Fresh ``MockVectorStore`` instance."""
    return MockVectorStore()


@pytest.fixture()
def mock_structured_db() -> MockStructuredDB:
    """Fresh ``MockStructuredDB`` instance."""
    return MockStructuredDB()


@pytest.fixture()
def mock_llm() -> MockLLM:
    """Fresh ``MockLLM`` instance with empty response queues."""
    return MockLLM()


@pytest.fixture()
def mock_embedding() -> MockEmbedding:
    """Fresh ``MockEmbedding`` instance (768-dim zero vectors)."""
    return MockEmbedding()


@pytest.fixture()
def test_config() -> ExcelProcessorConfig:
    """``ExcelProcessorConfig`` pre-set with test-friendly values."""
    return ExcelProcessorConfig(
        tenant_id="test_tenant",
        log_sample_data=True,
    )


# ---------------------------------------------------------------------------
# Session-scoped .xlsx Fixture Generators
# ---------------------------------------------------------------------------

_XLSX_TMP_DIR: tempfile.TemporaryDirectory | None = None


def _xlsx_dir() -> pathlib.Path:
    """Lazily create a session-wide temp directory for generated .xlsx files."""
    global _XLSX_TMP_DIR  # noqa: PLW0603
    if _XLSX_TMP_DIR is None:
        _XLSX_TMP_DIR = tempfile.TemporaryDirectory(prefix="ingestkit_test_xlsx_")
    return pathlib.Path(_XLSX_TMP_DIR.name)


@pytest.fixture(scope="session")
def type_a_simple_xlsx() -> pathlib.Path:
    """Generate a simple tabular .xlsx: 3 columns (ID, Name, Value), 20 rows."""
    path = _xlsx_dir() / "type_a_simple.xlsx"
    if path.exists():
        return path
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ID", "Name", "Value"])
    for i in range(1, 21):
        ws.append([i, f"Item-{i}", round(i * 1.5, 2)])
    wb.save(path)
    return path


@pytest.fixture(scope="session")
def type_b_checklist_xlsx() -> pathlib.Path:
    """Generate a formatted-document .xlsx with merged header and checklist items."""
    path = _xlsx_dir() / "type_b_checklist.xlsx"
    if path.exists():
        return path
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checklist"
    # Merged header spanning A1:C1
    ws.merge_cells("A1:C1")
    ws["A1"] = "Compliance Checklist"
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
    # Checklist items (text-heavy, no numeric columns)
    items = [
        "Verify fire extinguisher inspection dates",
        "Check emergency exit signage illumination",
        "Review first-aid kit supply levels",
        "Confirm safety data sheets are up to date",
        "Inspect PPE storage conditions",
        "Validate chemical labeling compliance",
        "Test emergency alarm system functionality",
        "Audit workplace ergonomic assessments",
        "Review incident report filing procedures",
        "Check electrical panel access clearance",
    ]
    for idx, item in enumerate(items, start=2):
        ws.cell(row=idx, column=1, value=f"{idx - 1}.")
        ws.cell(row=idx, column=2, value=item)
        ws.cell(row=idx, column=3, value="Pending")
    wb.save(path)
    return path


@pytest.fixture(scope="session")
def type_c_hybrid_xlsx() -> pathlib.Path:
    """Generate a hybrid .xlsx: Sheet1 = 10-row table, Sheet2 = formatted text with merges."""
    path = _xlsx_dir() / "type_c_hybrid.xlsx"
    if path.exists():
        return path
    wb = openpyxl.Workbook()

    # Sheet 1 — tabular
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Date", "Product", "Quantity", "Revenue"])
    for i in range(1, 11):
        ws1.append([f"2025-01-{i:02d}", f"Widget-{i}", i * 10, round(i * 10 * 9.99, 2)])

    # Sheet 2 — formatted document
    ws2 = wb.create_sheet("Notes")
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Quarterly Review Notes"
    ws2["A1"].font = openpyxl.styles.Font(bold=True, size=14)
    notes = [
        "Overall performance exceeded targets by 12%.",
        "Widget-3 had a supply chain delay in week 2.",
        "Customer satisfaction scores improved Q-over-Q.",
        "Recommend increasing Widget-7 production capacity.",
    ]
    for idx, note in enumerate(notes, start=3):
        ws2.merge_cells(f"A{idx}:D{idx}")
        ws2.cell(row=idx, column=1, value=note)

    wb.save(path)
    return path


@pytest.fixture(scope="session")
def edge_empty_xlsx() -> pathlib.Path:
    """Generate an empty .xlsx workbook (no data in any cell)."""
    path = _xlsx_dir() / "edge_empty.xlsx"
    if path.exists():
        return path
    wb = openpyxl.Workbook()
    wb.save(path)
    return path


@pytest.fixture(scope="session")
def edge_chart_only_xlsx() -> pathlib.Path:
    """Generate a .xlsx with chart data and a chart sheet.

    openpyxl can create charts but not true chartsheet-only workbooks in all
    cases.  This fixture creates a data sheet + an embedded chart as a
    reasonable approximation.
    """
    path = _xlsx_dir() / "edge_chart_only.xlsx"
    if path.exists():
        return path
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ChartData"
    ws.append(["Category", "Value"])
    for i in range(1, 6):
        ws.append([f"Cat-{i}", i * 100])

    chart = BarChart()
    chart.title = "Sample Chart"
    data = Reference(ws, min_col=2, min_row=1, max_row=6)
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")

    wb.save(path)
    return path


@pytest.fixture(scope="session")
def edge_large_xlsx() -> pathlib.Path:
    """Generate a large .xlsx with 100,001 rows (exceeds default max_rows_in_memory).

    Uses write-only mode for speed and minimal memory footprint.
    """
    path = _xlsx_dir() / "edge_large.xlsx"
    if path.exists():
        return path
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("LargeData")
    ws.append(["ID", "Value"])
    for i in range(1, 100_002):
        ws.append([i, i * 0.1])
    wb.save(path)
    return path

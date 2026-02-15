"""ExcelCellExtractor: openpyxl cell value mapping.

Reads form field values from Excel files using cell address mappings
defined in the FormTemplate (spec section 7.3).
"""

from __future__ import annotations

import logging
import re
import threading
from datetime import datetime
from typing import Any

import openpyxl

from ingestkit_forms.config import FormProcessorConfig
from ingestkit_forms.errors import FormErrorCode, FormIngestException
from ingestkit_forms.models import (
    ExtractedField,
    FieldMapping,
    FieldType,
    FormTemplate,
)

logger = logging.getLogger("ingestkit_forms")

_REGEX_TIMEOUT_SECONDS = 1.0


def _regex_match_with_timeout(
    pattern: str,
    value: str,
    timeout: float = _REGEX_TIMEOUT_SECONDS,
) -> bool | None:
    """Match a regex pattern with timeout protection against ReDoS.

    Uses ``re.fullmatch`` for validation (must match entire string).

    Returns:
        True if matches, False if no match, None if timeout or invalid pattern.
    """
    result: list[bool | None] = []

    def _match() -> None:
        try:
            result.append(bool(re.fullmatch(pattern, value)))
        except re.error:
            result.append(None)

    thread = threading.Thread(target=_match, daemon=True)
    thread.start()
    thread.join(timeout=timeout)

    if not result:
        # Timeout occurred
        return None
    return result[0]


class ExcelCellExtractor:
    """Extracts form field values from Excel files using cell address mappings.

    Uses openpyxl to read cell values from designated addresses defined
    in the FormTemplate (spec section 7.3). Supports single cells, cell
    ranges, and merged cell resolution.
    """

    def __init__(self, config: FormProcessorConfig) -> None:
        self._config = config

    def extract(
        self,
        file_path: str,
        template: FormTemplate,
    ) -> list[ExtractedField]:
        """Extract form field values from an Excel file.

        Opens the workbook with ``data_only=True`` to read formula results
        and ``read_only=False`` to access merged cell information.

        Args:
            file_path: Path to the Excel file.
            template: Form template with field mappings.

        Returns:
            List of ExtractedField objects, one per template field with
            a cell_address.

        Raises:
            FormIngestException: If the workbook cannot be opened (corrupt file).
        """
        try:
            wb = openpyxl.load_workbook(
                file_path, read_only=False, data_only=True
            )
        except Exception as exc:
            raise FormIngestException(
                code=FormErrorCode.E_FORM_FILE_CORRUPT,
                message=f"Failed to open Excel workbook: {exc}",
                stage="excel_cell_extraction",
                recoverable=False,
            ) from exc

        try:
            results: list[ExtractedField] = []
            for field in template.fields:
                if field.cell_address is None:
                    continue
                extracted = self._extract_field(wb, field)
                results.append(extracted)
            return results
        finally:
            wb.close()

    def _extract_field(
        self,
        wb: openpyxl.Workbook,
        field: FieldMapping,
    ) -> ExtractedField:
        """Extract a single field from the workbook."""
        assert field.cell_address is not None  # noqa: S101

        warnings: list[str] = []

        # Resolve worksheet
        try:
            if field.cell_address.sheet_name is not None:
                ws = wb[field.cell_address.sheet_name]
            else:
                ws = wb.active
        except KeyError:
            return ExtractedField(
                field_id=field.field_id,
                field_name=field.field_name,
                field_label=field.field_label,
                field_type=field.field_type,
                value=None,
                raw_value=None,
                confidence=0.0,
                extraction_method="cell_mapping",
                validation_passed=None,
                warnings=[
                    FormErrorCode.E_FORM_EXTRACTION_FAILED.value,
                    f"Sheet '{field.cell_address.sheet_name}' not found in workbook",
                ],
            )

        # Read cell value: range vs single cell
        cell_ref = field.cell_address.cell
        try:
            if ":" in cell_ref:
                raw_value, range_warnings = self._read_range(ws, cell_ref)
                warnings.extend(range_warnings)
            else:
                raw_value, was_merged = self._resolve_merged_cell(ws, cell_ref)
                if was_merged:
                    warnings.append(FormErrorCode.W_FORM_MERGED_CELL_RESOLVED.value)
        except Exception:
            return ExtractedField(
                field_id=field.field_id,
                field_name=field.field_name,
                field_label=field.field_label,
                field_type=field.field_type,
                value=None,
                raw_value=None,
                confidence=0.0,
                extraction_method="cell_mapping",
                validation_passed=None,
                warnings=[
                    FormErrorCode.E_FORM_EXTRACTION_FAILED.value,
                    f"Invalid cell address: {cell_ref}",
                ],
            )

        # Determine raw_value string for provenance
        raw_str = str(raw_value) if raw_value is not None else None

        # Handle empty cells
        if raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ""):
            if field.required:
                warnings.append(FormErrorCode.W_FORM_FIELD_MISSING_REQUIRED.value)
                value, validation_passed, val_warnings = self._validate_field_value(
                    None, field
                )
                warnings.extend(val_warnings)
                return ExtractedField(
                    field_id=field.field_id,
                    field_name=field.field_name,
                    field_label=field.field_label,
                    field_type=field.field_type,
                    value=None,
                    raw_value=raw_str,
                    confidence=0.0,
                    extraction_method="cell_mapping",
                    validation_passed=validation_passed,
                    warnings=warnings,
                )
            else:
                # Optional field, use default
                default = field.default_value
                value, validation_passed, val_warnings = self._validate_field_value(
                    default, field
                )
                warnings.extend(val_warnings)
                return ExtractedField(
                    field_id=field.field_id,
                    field_name=field.field_name,
                    field_label=field.field_label,
                    field_type=field.field_type,
                    value=default,
                    raw_value=raw_str,
                    confidence=0.95,
                    extraction_method="cell_mapping",
                    validation_passed=validation_passed,
                    warnings=warnings,
                )

        # Type coercion
        coerced, coercion_warnings = self._coerce_value(raw_value, field.field_type)
        warnings.extend(coercion_warnings)

        # Validation
        value, validation_passed, val_warnings = self._validate_field_value(
            coerced, field
        )
        warnings.extend(val_warnings)

        # PII-safe logging
        if self._config.log_sample_data and value is not None:
            logger.debug(
                "Excel cell value for field '%s': %s", field.field_name, value
            )

        return ExtractedField(
            field_id=field.field_id,
            field_name=field.field_name,
            field_label=field.field_label,
            field_type=field.field_type,
            value=value,
            raw_value=raw_str,
            confidence=0.95,
            extraction_method="cell_mapping",
            validation_passed=validation_passed,
            warnings=warnings,
        )

    def _resolve_merged_cell(
        self, ws: Any, cell_coord: str
    ) -> tuple[Any, bool]:
        """Check if a cell is in a merged range; if so, read the top-left cell.

        Returns:
            (value, was_merged) -- the cell value and whether it was resolved
            from a merged range.
        """
        for merged_range in ws.merged_cells.ranges:
            if cell_coord in merged_range:
                # Read from top-left cell of the merged range
                top_left = merged_range.min_row, merged_range.min_col
                value = ws.cell(row=top_left[0], column=top_left[1]).value
                return value, True

        return ws[cell_coord].value, False

    def _read_range(
        self, ws: Any, range_str: str
    ) -> tuple[str, list[str]]:
        """Read all cells in a range, join non-empty values with newline.

        Ranges always produce TEXT values before coercion (per PLAN-CHECK
        correction). The joined string is then passed to ``_coerce_value``.

        Returns:
            (joined_value, warnings)
        """
        warnings: list[str] = []
        values: list[str] = []

        for row in ws[range_str]:
            for cell in row:
                if cell.value is not None:
                    cell_str = str(cell.value).strip()
                    if cell_str:
                        values.append(cell_str)

        joined = "\n".join(values) if values else ""
        return joined, warnings

    @staticmethod
    def _coerce_value(
        raw_value: Any, field_type: FieldType
    ) -> tuple[Any, list[str]]:
        """Type coercion per spec 7.3d.

        Args:
            raw_value: The raw cell value from openpyxl.
            field_type: Expected data type.

        Returns:
            (coerced_value, warnings)
        """
        warnings: list[str] = []

        if field_type == FieldType.NUMBER:
            if isinstance(raw_value, (int, float)) and not isinstance(raw_value, bool):
                return float(raw_value), warnings
            try:
                return float(raw_value), warnings
            except (ValueError, TypeError):
                warnings.append(FormErrorCode.W_FORM_FIELD_TYPE_COERCION.value)
                return None, warnings

        if field_type == FieldType.DATE:
            if isinstance(raw_value, datetime):
                return raw_value.isoformat(), warnings
            # String fallback: try common formats to validate, return as-is
            if isinstance(raw_value, str):
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S"):
                    try:
                        datetime.strptime(raw_value, fmt)
                        return raw_value, warnings
                    except ValueError:
                        continue
            warnings.append(FormErrorCode.W_FORM_FIELD_TYPE_COERCION.value)
            return None, warnings

        if field_type == FieldType.CHECKBOX:
            # True values: "X", "x", "Yes", "TRUE", 1, True
            if isinstance(raw_value, bool):
                return raw_value, warnings
            if isinstance(raw_value, (int, float)):
                return raw_value == 1, warnings
            if isinstance(raw_value, str):
                return raw_value.strip().lower() in {"x", "yes", "true", "1"}, warnings
            return False, warnings

        # TEXT, DROPDOWN, RADIO, SIGNATURE -- all coerce to str
        return str(raw_value).strip(), warnings

    @staticmethod
    def _validate_field_value(
        value: Any,
        field: FieldMapping,
    ) -> tuple[Any, bool | None, list[str]]:
        """Validate an extracted value against field constraints.

        Uses ReDoS-protected regex matching with a 1-second timeout,
        consistent with NativePDFExtractor and OCROverlayExtractor.

        Returns:
            (value, validation_passed, warnings)
        """
        if field.validation_pattern is None:
            return (value, None, [])

        if value is None:
            return (None, None, [])

        str_value = str(value)
        match_result = _regex_match_with_timeout(
            field.validation_pattern, str_value
        )

        if match_result is None:
            # Timeout or invalid pattern
            return (
                value,
                False,
                [FormErrorCode.W_FORM_FIELD_VALIDATION_FAILED.value],
            )
        if match_result:
            return (value, True, [])
        return (
            value,
            False,
            [FormErrorCode.W_FORM_FIELD_VALIDATION_FAILED.value],
        )

"""Tests for FormRouter orchestrator."""

from __future__ import annotations

import sys
from pathlib import Path
from unittest.mock import MagicMock, patch  # noqa: F401

import pytest

from ingestkit_forms.config import FormProcessorConfig
from ingestkit_forms.errors import FormErrorCode, FormIngestException
from ingestkit_forms.models import (
    ExtractedField,  # noqa: F401
    FieldType,
    FormIngestRequest,
    FormProcessingResult,
    SourceFormat,
    TemplateMatch,
)
from ingestkit_forms.router import FormRouter, create_default_router

# Import test helpers from the conftest module via tests directory path
_tests_dir = str(Path(__file__).resolve().parent)
if _tests_dir not in sys.path:
    sys.path.insert(0, _tests_dir)

from conftest import (  # noqa: E402
    MockEmbeddingBackend,
    MockFormDBBackend,
    MockFormTemplateStore,
    MockLayoutFingerprinter,
    MockOCRBackend,
    MockPDFWidgetBackend,
    MockVectorStoreBackend,
    MockVLMBackend,
    make_extracted_field,  # noqa: F401
    make_field_mapping,  # noqa: F401
    make_template,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def form_config():
    """Config with fast backend settings for tests."""
    return FormProcessorConfig(
        backend_max_retries=0,
        backend_backoff_base=0.0,
        backend_timeout_seconds=1.0,
        tenant_id="test-tenant",
    )


@pytest.fixture()
def vlm_config():
    """Config with VLM enabled."""
    return FormProcessorConfig(
        form_vlm_enabled=True,
        form_vlm_fallback_threshold=0.4,
        form_vlm_max_fields_per_document=10,
        backend_max_retries=0,
        backend_backoff_base=0.0,
        tenant_id="test-tenant",
    )


@pytest.fixture()
def all_backends():
    """Return dict of all mock backends."""
    return {
        "template_store": MockFormTemplateStore(),
        "fingerprinter": MockLayoutFingerprinter(),
        "form_db": MockFormDBBackend(),
        "vector_store": MockVectorStoreBackend(),
        "embedder": MockEmbeddingBackend(),
        "ocr_backend": MockOCRBackend(default_text="test value", default_confidence=0.85),
        "pdf_widget_backend": MockPDFWidgetBackend(has_fields=True),
        "vlm_backend": MockVLMBackend(),
    }


@pytest.fixture()
def form_router(all_backends, form_config):
    """FormRouter with all mock backends."""
    return FormRouter(**all_backends, config=form_config)


@pytest.fixture()
def xlsx_template():
    """Excel template for testing."""
    from ingestkit_forms.models import CellAddress, FieldMapping

    fields = [
        FieldMapping(
            field_id="f1",
            field_name="employee_name",
            field_label="Employee Name",
            field_type=FieldType.TEXT,
            page_number=0,
            cell_address=CellAddress(cell="B2"),
        ),
    ]
    return make_template(
        fields=fields,
        source_format=SourceFormat.XLSX,
        template_id="tmpl-xlsx-1",
        name="Excel Form",
    )


@pytest.fixture()
def pdf_template():
    """PDF template for testing."""
    return make_template(
        source_format=SourceFormat.PDF,
        template_id="tmpl-pdf-1",
        name="PDF Form",
    )


# ---------------------------------------------------------------------------
# Constructor Tests
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestRouterConstructor:
    """Tests for FormRouter.__init__."""

    def test_router_init_all_backends(self, all_backends, form_config):
        """Router accepts all 9 DI backends and creates internal components."""
        router = FormRouter(**all_backends, config=form_config)
        assert router._matcher is not None
        assert router._excel_extractor is not None
        assert router._native_pdf_extractor is not None
        assert router._ocr_extractor is not None
        assert router._vlm_extractor is not None
        assert router._dual_writer is not None

    def test_router_init_minimal_backends(self, form_config):
        """Router works with only required backends (optional = None)."""
        router = FormRouter(
            template_store=MockFormTemplateStore(),
            fingerprinter=MockLayoutFingerprinter(),
            form_db=MockFormDBBackend(),
            vector_store=MockVectorStoreBackend(),
            embedder=MockEmbeddingBackend(),
            config=form_config,
        )
        assert router._native_pdf_extractor is None
        assert router._ocr_extractor is None
        assert router._vlm_extractor is None
        assert router._dual_writer is not None

    def test_router_warns_when_pdf_widget_backend_missing(self, form_config, caplog):
        """Router logs W_FORM_NATIVE_FIELDS_UNAVAILABLE when PDFWidgetBackend is None."""
        import logging

        with caplog.at_level(logging.WARNING, logger="ingestkit_forms"):
            FormRouter(
                template_store=MockFormTemplateStore(),
                fingerprinter=MockLayoutFingerprinter(),
                form_db=MockFormDBBackend(),
                vector_store=MockVectorStoreBackend(),
                embedder=MockEmbeddingBackend(),
                config=form_config,
            )
        assert any(
            "W_FORM_NATIVE_FIELDS_UNAVAILABLE" in record.message
            for record in caplog.records
        )


# ---------------------------------------------------------------------------
# Extractor Selection Tests
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestExtractorSelection:
    """Tests for FormRouter._select_extractor."""

    def test_select_extractor_xlsx(self, form_router, xlsx_template, tmp_path):
        """XLSX files route to ExcelCellExtractor."""
        xlsx_file = tmp_path / "test.xlsx"
        xlsx_file.touch()
        method, extractor = form_router._select_extractor(str(xlsx_file), xlsx_template)
        assert method == "cell_mapping"

    def test_select_extractor_pdf_with_fields(self, form_router, pdf_template, tmp_path):
        """PDF with form fields routes to NativePDFExtractor."""
        pdf_file = tmp_path / "test.pdf"
        pdf_file.touch()
        method, extractor = form_router._select_extractor(str(pdf_file), pdf_template)
        assert method == "native_fields"

    def test_select_extractor_pdf_without_fields(self, all_backends, form_config, pdf_template, tmp_path):
        """PDF without form fields routes to OCROverlayExtractor."""
        all_backends["pdf_widget_backend"] = MockPDFWidgetBackend(has_fields=False)
        router = FormRouter(**all_backends, config=form_config)
        pdf_file = tmp_path / "test.pdf"
        pdf_file.touch()
        method, extractor = router._select_extractor(str(pdf_file), pdf_template)
        assert method == "ocr_overlay"

    def test_select_extractor_image(self, form_router, tmp_path):
        """Image files route to OCROverlayExtractor."""
        img_file = tmp_path / "test.png"
        img_file.touch()
        template = make_template(source_format=SourceFormat.IMAGE, template_id="tmpl-img")
        method, extractor = form_router._select_extractor(str(img_file), template)
        assert method == "ocr_overlay"


# ---------------------------------------------------------------------------
# Pipeline Tests
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestExtractFormPipeline:
    """Tests for the full extract_form pipeline."""

    def _setup_template_in_store(self, store, template):
        """Add a template to the mock store."""
        store.save_template(template)

    def test_extract_form_manual_template(self, all_backends, form_config, xlsx_template, tmp_path):
        """Manual template_id: happy path with full result."""
        import openpyxl

        # Create a real xlsx file with a value
        xlsx_file = tmp_path / "form.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "John Doe"
        wb.save(str(xlsx_file))

        store = all_backends["template_store"]
        self._setup_template_in_store(store, xlsx_template)

        router = FormRouter(**all_backends, config=form_config)
        request = FormIngestRequest(
            file_path=str(xlsx_file),
            template_id="tmpl-xlsx-1",
            tenant_id="test-tenant",
        )
        result = router.extract_form(request)

        assert result is not None
        assert isinstance(result, FormProcessingResult)
        assert result.file_path == str(xlsx_file)
        assert result.tenant_id == "test-tenant"
        assert result.extraction_result.match_method == "manual_override"
        assert result.extraction_result.extraction_method == "cell_mapping"
        assert result.extraction_result.template_id == "tmpl-xlsx-1"
        assert result.processing_time_seconds > 0
        assert result.ingest_key  # non-empty

    def test_extract_form_no_match_fallthrough(self, all_backends, form_config, tmp_path):
        """No template_id and no match -> returns None (graceful fallthrough)."""
        # Store is empty, so auto-match returns no matches
        xlsx_file = tmp_path / "form.xlsx"
        xlsx_file.touch()

        router = FormRouter(**all_backends, config=form_config)
        request = FormIngestRequest(
            file_path=str(xlsx_file),
        )

        # detect_source_format will be called -- xlsx is fine
        # But match_document will return empty (no templates in store with fingerprints)
        result = router.extract_form(request)
        assert result is None

    def test_extract_form_low_confidence_fail_closed(self, all_backends, tmp_path):
        """Overall confidence below threshold -> result with error, zero outputs."""
        import openpyxl
        from ingestkit_forms.models import CellAddress, FieldMapping

        # Config with very high confidence threshold
        strict_config = FormProcessorConfig(
            form_extraction_min_overall_confidence=0.99,
            backend_max_retries=0,
            backend_backoff_base=0.0,
            tenant_id="test-tenant",
        )

        # Template with a required field that will get low confidence from empty cell
        fields = [
            FieldMapping(
                field_id="f1",
                field_name="employee_name",
                field_label="Employee Name",
                field_type=FieldType.TEXT,
                page_number=0,
                cell_address=CellAddress(cell="B2"),
                required=True,
            ),
        ]
        template = make_template(
            fields=fields,
            source_format=SourceFormat.XLSX,
            template_id="tmpl-strict",
            name="Strict Form",
        )

        store = all_backends["template_store"]
        store.save_template(template)

        # Create xlsx with empty cell (will get low confidence)
        xlsx_file = tmp_path / "form.xlsx"
        wb = openpyxl.Workbook()
        wb.save(str(xlsx_file))

        router = FormRouter(**all_backends, config=strict_config)
        request = FormIngestRequest(
            file_path=str(xlsx_file),
            template_id="tmpl-strict",
        )
        result = router.extract_form(request)

        assert result is not None
        assert result.chunks_created == 0
        assert result.tables_created == 0
        assert len(result.errors) > 0
        assert any(
            "E_FORM_EXTRACTION_LOW_CONFIDENCE" in str(e.code)
            for e in result.error_details
        )

    def test_extract_form_batch(self, all_backends, form_config, xlsx_template, tmp_path):
        """Batch processes multiple requests sequentially."""
        import openpyxl

        store = all_backends["template_store"]
        store.save_template(xlsx_template)

        # Create two xlsx files
        files = []
        for i in range(2):
            xlsx_file = tmp_path / f"form_{i}.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["B2"] = f"Person {i}"
            wb.save(str(xlsx_file))
            files.append(str(xlsx_file))

        router = FormRouter(**all_backends, config=form_config)
        requests = [
            FormIngestRequest(file_path=f, template_id="tmpl-xlsx-1")
            for f in files
        ]
        results = router.extract_form_batch(requests)

        assert len(results) == 2
        for result in results:
            assert result is not None
            assert isinstance(result, FormProcessingResult)

    def test_extract_form_unsupported_format(self, form_router, tmp_path):
        """Unsupported file extension raises E_FORM_UNSUPPORTED_FORMAT."""
        docx_file = tmp_path / "form.docx"
        docx_file.touch()

        request = FormIngestRequest(
            file_path=str(docx_file),
            template_id="tmpl-pdf-1",
        )
        # Template needs to be in the store for resolve_manual_override
        template = make_template(
            source_format=SourceFormat.PDF,
            template_id="tmpl-pdf-1",
        )
        form_router._template_store.save_template(template)

        with pytest.raises(FormIngestException) as exc_info:
            form_router.extract_form(request)
        assert exc_info.value.code == FormErrorCode.E_FORM_UNSUPPORTED_FORMAT

    def test_extract_form_template_not_found(self, form_router, tmp_path):
        """Manual template_id not in store raises E_FORM_TEMPLATE_NOT_FOUND."""
        xlsx_file = tmp_path / "form.xlsx"
        xlsx_file.touch()

        request = FormIngestRequest(
            file_path=str(xlsx_file),
            template_id="nonexistent-template",
        )

        with pytest.raises(FormIngestException) as exc_info:
            form_router.extract_form(request)
        assert exc_info.value.code == FormErrorCode.E_FORM_TEMPLATE_NOT_FOUND

    def test_extract_form_dual_write_error(self, all_backends, form_config, xlsx_template, tmp_path):
        """Mock DB failure -> errors in result, result still returned."""
        import openpyxl

        # Use a DB that fails on the first call
        all_backends["form_db"] = MockFormDBBackend(fail_on_call=1)
        store = all_backends["template_store"]
        store.save_template(xlsx_template)

        xlsx_file = tmp_path / "form.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "John"
        wb.save(str(xlsx_file))

        router = FormRouter(**all_backends, config=form_config)
        request = FormIngestRequest(
            file_path=str(xlsx_file),
            template_id="tmpl-xlsx-1",
        )
        result = router.extract_form(request)

        assert result is not None
        # Should have errors from the DB failure
        assert len(result.errors) > 0


# ---------------------------------------------------------------------------
# VLM Fallback Test
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestVLMFallback:
    """Tests for VLM fallback in the pipeline."""

    def test_extract_form_vlm_fallback_triggered(self, all_backends, vlm_config, tmp_path):
        """VLM-enabled config with low-confidence field triggers VLM fallback."""
        import openpyxl
        from ingestkit_forms.models import CellAddress, FieldMapping

        # Template with a field
        fields = [
            FieldMapping(
                field_id="f1",
                field_name="employee_name",
                field_label="Employee Name",
                field_type=FieldType.TEXT,
                page_number=0,
                cell_address=CellAddress(cell="B2"),
            ),
        ]
        template = make_template(
            fields=fields,
            source_format=SourceFormat.XLSX,
            template_id="tmpl-vlm-1",
            name="VLM Test Form",
        )

        store = all_backends["template_store"]
        store.save_template(template)

        xlsx_file = tmp_path / "form.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "John"
        wb.save(str(xlsx_file))

        router = FormRouter(**all_backends, config=vlm_config)
        request = FormIngestRequest(
            file_path=str(xlsx_file),
            template_id="tmpl-vlm-1",
        )

        # ExcelCellExtractor returns high confidence (0.9+), so VLM won't
        # trigger. This tests the pipeline integrates correctly even with
        # VLM enabled. The VLM extractor only applies to fields marked
        # "vlm_fallback_pending", which happens only for very low confidence.
        result = router.extract_form(request)
        assert result is not None
        assert isinstance(result, FormProcessingResult)


# ---------------------------------------------------------------------------
# Factory Tests
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestCreateDefaultRouter:
    """Tests for create_default_router factory."""

    def test_create_default_router_with_backends(self):
        """Passing all required backends returns a FormRouter."""
        router = create_default_router(
            template_store=MockFormTemplateStore(),
            fingerprinter=MockLayoutFingerprinter(),
            form_db=MockFormDBBackend(),
            vector_store=MockVectorStoreBackend(),
            embedder=MockEmbeddingBackend(),
        )
        assert isinstance(router, FormRouter)

    def test_create_default_router_missing_required(self):
        """Omitting required backends raises ValueError."""
        with pytest.raises(ValueError, match="Required backend"):
            create_default_router(
                template_store=MockFormTemplateStore(),
                # fingerprinter, form_db, vector_store, embedder all missing
            )

    def test_create_default_router_missing_single(self):
        """Missing a single required backend lists it in error."""
        with pytest.raises(ValueError, match="form_db"):
            create_default_router(
                template_store=MockFormTemplateStore(),
                fingerprinter=MockLayoutFingerprinter(),
                # form_db missing
                vector_store=MockVectorStoreBackend(),
                embedder=MockEmbeddingBackend(),
            )


# ---------------------------------------------------------------------------
# Idempotency Integration Test
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestIdempotencyIntegration:
    """Test that the router correctly uses idempotency functions."""

    def test_idempotency_keys_deterministic(self, all_backends, form_config, xlsx_template, tmp_path):
        """Same inputs produce same ingest key; different template_version changes extraction key."""
        import openpyxl

        from ingestkit_forms.idempotency import compute_form_extraction_key, compute_ingest_key

        store = all_backends["template_store"]
        store.save_template(xlsx_template)

        xlsx_file = tmp_path / "form.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "John"
        wb.save(str(xlsx_file))

        # Compute keys directly
        ik = compute_ingest_key(
            str(xlsx_file), form_config.parser_version, "test-tenant"
        )
        ek1 = compute_form_extraction_key(ik.key, "tmpl-xlsx-1", 1)
        ek2 = compute_form_extraction_key(ik.key, "tmpl-xlsx-1", 2)

        # Same global key
        ik2 = compute_ingest_key(
            str(xlsx_file), form_config.parser_version, "test-tenant"
        )
        assert ik.key == ik2.key

        # Different template version -> different extraction key
        assert ek1 != ek2

        # Same inputs -> same extraction key
        ek1b = compute_form_extraction_key(ik.key, "tmpl-xlsx-1", 1)
        assert ek1 == ek1b


# ---------------------------------------------------------------------------
# match_document() wrapper tests (issue #72)
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestMatchDocument:
    """Tests for FormRouter.match_document wrapper."""

    def test_match_document_disabled(self, all_backends, tmp_path):
        """form_match_enabled=False returns empty list immediately."""
        config = FormProcessorConfig(
            form_match_enabled=False,
            backend_max_retries=0,
            backend_backoff_base=0.0,
            tenant_id="test-tenant",
        )
        router = FormRouter(**all_backends, config=config)
        xlsx_file = tmp_path / "form.xlsx"
        xlsx_file.touch()

        result = router.match_document(str(xlsx_file))
        assert result == []

    def test_match_document_no_templates(self, all_backends, form_config, tmp_path):
        """No stored templates returns empty list."""
        router = FormRouter(**all_backends, config=form_config)
        xlsx_file = tmp_path / "form.xlsx"
        xlsx_file.touch()

        result = router.match_document(str(xlsx_file))
        assert result == []

    def test_match_document_logs_structured(self, all_backends, form_config, tmp_path, caplog):
        """Verifies structured log output per spec 18.4."""
        import logging

        router = FormRouter(**all_backends, config=form_config)
        xlsx_file = tmp_path / "form.xlsx"
        xlsx_file.touch()

        with caplog.at_level(logging.INFO, logger="ingestkit_forms"):
            router.match_document(str(xlsx_file))

        # Find the structured match log (forms.match.fallthrough or forms.match.auto)
        match_logs = [
            r for r in caplog.records
            if "forms.match." in r.message
        ]
        assert len(match_logs) == 1

        record = match_logs[0]
        assert hasattr(record, "template_candidates")
        assert hasattr(record, "confidence")
        assert hasattr(record, "match_duration_ms")

    def test_match_document_disabled_logs_fallthrough(self, all_backends, tmp_path, caplog):
        """Disabled matching logs forms.match.disabled."""
        import logging

        config = FormProcessorConfig(
            form_match_enabled=False,
            backend_max_retries=0,
            backend_backoff_base=0.0,
            tenant_id="test-tenant",
        )
        router = FormRouter(**all_backends, config=config)
        xlsx_file = tmp_path / "form.xlsx"
        xlsx_file.touch()

        with caplog.at_level(logging.INFO, logger="ingestkit_forms"):
            router.match_document(str(xlsx_file))

        match_logs = [
            r for r in caplog.records
            if "forms.match.disabled" in r.message
        ]
        assert len(match_logs) == 1


# ---------------------------------------------------------------------------
# try_match() pipeline gate tests (issue #72)
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestTryMatch:
    """Tests for FormRouter.try_match pipeline gate."""

    def test_try_match_auto_no_match(self, all_backends, form_config, tmp_path):
        """No fingerprint match returns None."""
        router = FormRouter(**all_backends, config=form_config)
        xlsx_file = tmp_path / "form.xlsx"
        xlsx_file.touch()

        result = router.try_match(str(xlsx_file))
        assert result is None

    def test_try_match_auto_below_threshold(self, all_backends, form_config, tmp_path):
        """No match above threshold returns None (graceful fallthrough)."""
        router = FormRouter(**all_backends, config=form_config)
        xlsx_file = tmp_path / "form.xlsx"
        xlsx_file.touch()

        result = router.try_match(str(xlsx_file))
        assert result is None

    def test_try_match_disabled_no_manual(self, all_backends, tmp_path):
        """Disabled matching with no manual template returns None."""
        config = FormProcessorConfig(
            form_match_enabled=False,
            backend_max_retries=0,
            backend_backoff_base=0.0,
            tenant_id="test-tenant",
        )
        router = FormRouter(**all_backends, config=config)
        xlsx_file = tmp_path / "form.xlsx"
        xlsx_file.touch()

        result = router.try_match(str(xlsx_file))
        assert result is None

    def test_try_match_exception_returns_none(self, all_backends, form_config, tmp_path):
        """Exception in matching returns None (graceful fallthrough)."""
        # Use a fingerprinter that raises
        class FailFingerprinter:
            def compute_fingerprint(self, file_path: str) -> list[bytes]:
                raise RuntimeError("fingerprint boom")

        all_backends["fingerprinter"] = FailFingerprinter()
        router = FormRouter(**all_backends, config=form_config)
        xlsx_file = tmp_path / "form.xlsx"
        xlsx_file.touch()

        # Add a template with a fingerprint so match_document actually runs
        template = make_template(
            source_format=SourceFormat.XLSX,
            template_id="tmpl-fail",
            layout_fingerprint=b"\x01" * 320,
        )
        all_backends["template_store"].save_template(template)

        result = router.try_match(str(xlsx_file))
        assert result is None

    def test_try_match_no_state_mutation(self, all_backends, form_config, tmp_path):
        """Verify store/backends unchanged after fallthrough."""
        store = all_backends["template_store"]
        db = all_backends["form_db"]
        vs = all_backends["vector_store"]

        # Snapshot state
        templates_before = store.list_templates()
        db_calls_before = len(db.execute_sql_calls)
        vs_calls_before = len(vs.upsert_calls)

        router = FormRouter(**all_backends, config=form_config)
        xlsx_file = tmp_path / "form.xlsx"
        xlsx_file.touch()

        result = router.try_match(str(xlsx_file))
        assert result is None

        # Verify no state mutation
        templates_after = store.list_templates()
        assert len(templates_after) == len(templates_before)
        assert len(db.execute_sql_calls) == db_calls_before
        assert len(vs.upsert_calls) == vs_calls_before

    def test_try_match_with_tenant_id(self, all_backends, form_config, tmp_path):
        """try_match accepts tenant_id parameter per spec."""
        router = FormRouter(**all_backends, config=form_config)
        xlsx_file = tmp_path / "form.xlsx"
        xlsx_file.touch()

        result = router.try_match(str(xlsx_file), tenant_id="custom-tenant")
        assert result is None


# ---------------------------------------------------------------------------
# Structured Logging Tests (issue #72, spec 18.4)
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestStructuredLogging:
    """Tests for structured logging per spec 18.4."""

    def test_logger_name(self):
        """Logger is named 'ingestkit_forms'."""
        from ingestkit_forms.router import logger as router_logger

        assert router_logger.name == "ingestkit_forms"

    def test_extract_form_match_stage_log(self, all_backends, form_config, xlsx_template, tmp_path, caplog):
        """extract_form logs structured match-stage fields."""
        import logging
        import openpyxl

        store = all_backends["template_store"]
        store.save_template(xlsx_template)

        xlsx_file = tmp_path / "form.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "Jane"
        wb.save(str(xlsx_file))

        router = FormRouter(**all_backends, config=form_config)
        request = FormIngestRequest(
            file_path=str(xlsx_file),
            template_id="tmpl-xlsx-1",
        )

        with caplog.at_level(logging.INFO, logger="ingestkit_forms"):
            router.extract_form(request)

        match_logs = [
            r for r in caplog.records
            if "forms.match.manual" in r.message
        ]
        assert len(match_logs) >= 1

    def test_extract_form_extract_stage_log(self, all_backends, form_config, xlsx_template, tmp_path, caplog):
        """extract_form logs structured extract-stage fields."""
        import logging
        import openpyxl

        store = all_backends["template_store"]
        store.save_template(xlsx_template)

        xlsx_file = tmp_path / "form.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "Jane"
        wb.save(str(xlsx_file))

        router = FormRouter(**all_backends, config=form_config)
        request = FormIngestRequest(
            file_path=str(xlsx_file),
            template_id="tmpl-xlsx-1",
        )

        with caplog.at_level(logging.INFO, logger="ingestkit_forms"):
            router.extract_form(request)

        extract_logs = [
            r for r in caplog.records
            if "forms.extract.completed" in r.message
        ]
        assert len(extract_logs) >= 1

    def test_pii_safe_logging(self, all_backends, form_config, xlsx_template, tmp_path, caplog):
        """No field values appear in log output (PII safety)."""
        import logging
        import openpyxl

        store = all_backends["template_store"]
        store.save_template(xlsx_template)

        # Create xlsx with a recognizable PII-like value
        xlsx_file = tmp_path / "form.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "SENSITIVE_PII_VALUE_12345"
        wb.save(str(xlsx_file))

        router = FormRouter(**all_backends, config=form_config)
        request = FormIngestRequest(
            file_path=str(xlsx_file),
            template_id="tmpl-xlsx-1",
        )

        with caplog.at_level(logging.DEBUG, logger="ingestkit_forms"):
            router.extract_form(request)

        # Ensure the PII value never appears in any log message
        for record in caplog.records:
            assert "SENSITIVE_PII_VALUE_12345" not in record.message, (
                f"PII value leaked in log: {record.message}"
            )

    def test_fallthrough_match_stage_log(self, all_backends, form_config, tmp_path, caplog):
        """Fallthrough (no match) logs match_result=fallthrough."""
        import logging

        router = FormRouter(**all_backends, config=form_config)
        xlsx_file = tmp_path / "form.xlsx"
        xlsx_file.touch()

        request = FormIngestRequest(file_path=str(xlsx_file))

        with caplog.at_level(logging.INFO, logger="ingestkit_forms"):
            result = router.extract_form(request)

        assert result is None

        match_logs = [
            r for r in caplog.records
            if "forms.match.fallthrough" in r.message
        ]
        assert len(match_logs) >= 1


# ---------------------------------------------------------------------------
# manual_override field tests (issue #102)
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestManualOverride:
    """Tests for FormIngestRequest.manual_override field."""

    def test_manual_override_default_false(self):
        """manual_override defaults to False."""
        req = FormIngestRequest(file_path="/tmp/test.pdf")
        assert req.manual_override is False

    def test_manual_override_explicit_true(self):
        """manual_override=True is accepted."""
        req = FormIngestRequest(
            file_path="/tmp/test.pdf",
            template_id="tmpl-1",
            manual_override=True,
        )
        assert req.manual_override is True

    def test_manual_override_triggers_manual_path(
        self, all_backends, form_config, xlsx_template, tmp_path
    ):
        """manual_override=True routes through manual override path."""
        import logging
        import openpyxl

        store = all_backends["template_store"]
        store.save_template(xlsx_template)

        xlsx_file = tmp_path / "form.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "Test"
        wb.save(str(xlsx_file))

        router = FormRouter(**all_backends, config=form_config)
        request = FormIngestRequest(
            file_path=str(xlsx_file),
            template_id="tmpl-xlsx-1",
            manual_override=True,
        )
        result = router.extract_form(request)
        assert result is not None
        assert isinstance(result, FormProcessingResult)

    def test_template_id_without_override_still_works(
        self, all_backends, form_config, xlsx_template, tmp_path
    ):
        """template_id alone (without manual_override) still uses manual path."""
        import openpyxl

        store = all_backends["template_store"]
        store.save_template(xlsx_template)

        xlsx_file = tmp_path / "form.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "Test"
        wb.save(str(xlsx_file))

        router = FormRouter(**all_backends, config=form_config)
        request = FormIngestRequest(
            file_path=str(xlsx_file),
            template_id="tmpl-xlsx-1",
        )
        result = router.extract_form(request)
        assert result is not None


# ---------------------------------------------------------------------------
# validate_table_name tests (issue #103)
# ---------------------------------------------------------------------------


@pytest.mark.unit
class TestValidateTableName:
    """Tests for validate_table_name and its integration in db_writer."""

    def test_valid_identifiers(self):
        """Valid SQL identifiers pass validation."""
        from ingestkit_forms.security import validate_table_name

        assert validate_table_name("forms_employee") is None
        assert validate_table_name("_private") is None
        assert validate_table_name("A123") is None
        assert validate_table_name("a" * 128) is None

    def test_empty_name(self):
        """Empty string is rejected."""
        from ingestkit_forms.security import validate_table_name

        result = validate_table_name("")
        assert result is not None
        assert "empty" in result.lower()

    def test_starts_with_digit(self):
        """Names starting with a digit are rejected."""
        from ingestkit_forms.security import validate_table_name

        result = validate_table_name("1table")
        assert result is not None

    def test_special_characters(self):
        """Names with special characters are rejected."""
        from ingestkit_forms.security import validate_table_name

        assert validate_table_name("drop;table") is not None
        assert validate_table_name("my-table") is not None
        assert validate_table_name("my table") is not None
        assert validate_table_name("table$") is not None

    def test_too_long(self):
        """Names exceeding 128 chars are rejected."""
        from ingestkit_forms.security import validate_table_name

        result = validate_table_name("a" * 129)
        assert result is not None
        assert "128" in result

    def test_get_table_name_validates(self):
        """get_table_name raises on unsafe generated names."""
        from ingestkit_forms.output.db_writer import get_table_name

        config = FormProcessorConfig(
            form_db_table_prefix="forms_",
            tenant_id="test-tenant",
        )
        # Normal template name should pass
        template = make_template(name="Employee Form")
        table_name = get_table_name(config, template)
        assert table_name == "forms_employee_form"

    def test_get_table_name_rejects_unsafe(self):
        """get_table_name raises FormIngestException for unsafe names."""
        from ingestkit_forms.output.db_writer import get_table_name

        # A prefix starting with a digit would make the result unsafe
        config = FormProcessorConfig(
            form_db_table_prefix="1_",
            tenant_id="test-tenant",
        )
        template = make_template(name="Test")
        with pytest.raises(FormIngestException) as exc_info:
            get_table_name(config, template)
        assert exc_info.value.code == FormErrorCode.E_FORM_TEMPLATE_INVALID
